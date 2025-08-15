# invoice_timesheets_compiler_fixed_threads.py
# Thread-safe version (UI updates scheduled via .after)
import os
import time
import threading
import winsound
from datetime import datetime

import tkinter as tk
from tkinter import messagebox, ttk
import ttkbootstrap as tb
from ttkbootstrap.constants import *

from PyPDF2 import PdfMerger, PdfReader, PdfWriter
from PIL import Image, ImageOps
from docx2pdf import convert
import openpyxl

# ========== CONFIG ==========
main_folder = r"O:\ApTask\TDrive\FinTech LLC\Invoices\2025\Monthly"
excel_file  = r"C:\Users\HimalK\OneDrive - APTASK\Desktop\Aptask\Payroll\Automated email sheet\Emailexcel.xlsx"
log_folder  = r"C:\Users\HimalK\OneDrive - APTASK\Desktop\Aptask\Payroll\Timesheet_Invoice Merger\Logfile"
os.makedirs(log_folder, exist_ok=True)

clients_list = [
    "Aquila Energy", "BDR", "B Squared", "CFAIS", "Data Specialist",
    "HTS Workforce", "Schultz Controls", "Security 101", "VFS Fire", "Western Audio"
]

APP_TITLE = "Invoice and Timesheets Compiler"
THEME_BG  = "#2b6cb0"
PANEL_BG  = "#ffffff"
TEXT_FG   = "#000000"
ACCENT    = "#1e90ff"
# ============================


# Thread-safe logger that schedules UI updates on main thread
class StepLogger:
    def __init__(self, text_widget, week_str):
        self.text_widget = text_widget
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_week = week_str.replace("/", "-")
        self.path = os.path.join(log_folder, f"Log_Week_{safe_week}_{ts}.txt")

    def _write_file(self, line: str):
        try:
            with open(self.path, "a", encoding="utf-8") as f:
                f.write(line + "\n")
        except Exception:
            pass

    def log(self, msg, tag="info"):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        line = f"[{timestamp}] {msg}"
        # schedule UI update on main thread
        def ui_append():
            try:
                self.text_widget.configure(state="normal")
                if tag in ("error", "ok", "warn"):
                    self.text_widget.insert("end", line + "\n", (tag,))
                else:
                    self.text_widget.insert("end", line + "\n")
                self.text_widget.see("end")
                self.text_widget.configure(state="disabled")
            except Exception:
                # fallback: ignore UI error
                pass
        try:
            self.text_widget.after(0, ui_append)
        except Exception:
            # if widget cannot schedule, run inline (best-effort)
            ui_append()
        # always write to file (not UI)
        try:
            self._write_file(line)
        except Exception:
            pass


# ---------- Helpers ----------
def rotate_pdf_if_needed(src_pdf, dst_pdf):
    reader = PdfReader(src_pdf)
    writer = PdfWriter()
    for page in reader.pages:
        try:
            w = float(page.mediabox.width)
            h = float(page.mediabox.height)
            if w > h:
                page.rotate(90)
        except Exception:
            pass
        writer.add_page(page)
    with open(dst_pdf, "wb") as f:
        writer.write(f)


def image_to_pdf(image_path, out_pdf):
    image = Image.open(image_path)
    image = ImageOps.exif_transpose(image)
    image.convert("RGB").save(out_pdf)


def word_to_rotated_pdf(doc_path, out_pdf):
    temp_pdf = out_pdf.replace(".pdf", "_raw.pdf")
    convert(doc_path, temp_pdf)
    rotate_pdf_if_needed(temp_pdf, out_pdf)
    try:
        os.remove(temp_pdf)
    except Exception:
        pass


def find_week_folder(client_root, week_str):
    target = f"Week {week_str}"
    if not os.path.isdir(client_root):
        return None
    for month_name in sorted(os.listdir(client_root)):
        month_path = os.path.join(client_root, month_name)
        if not os.path.isdir(month_path):
            continue
        week_path = os.path.join(month_path, target)
        if os.path.isdir(week_path):
            return week_path
    return None


def list_raw_files(folder):
    files = []
    if not os.path.isdir(folder):
        return files
    for name in sorted(os.listdir(folder)):
        path = os.path.join(folder, name)
        if not os.path.isfile(path):
            continue
        ext = os.path.splitext(name)[1].lower()
        if ext in (".pdf", ".jpg", ".jpeg", ".png", ".docx", ".doc"):
            files.append(path)
    return files


def prepare_files_for_merge(folder, logger=None):
    invoice_candidate = None
    other_files = []
    for name in sorted(os.listdir(folder)):
        path = os.path.join(folder, name)
        if not os.path.isfile(path):
            continue
        lower = name.lower()
        root, ext = os.path.splitext(name)
        ext = ext.lower()
        if "invoice" in lower:
            invoice_candidate = path
            continue
        if ext in (".pdf", ".jpg", ".jpeg", ".png", ".docx", ".doc"):
            other_files.append(path)

    prepared = []
    invoice_final = None
    if invoice_candidate:
        try:
            r, ext = os.path.splitext(invoice_candidate)
            ext = ext.lower()
            if ext == ".pdf":
                invoice_final = os.path.join(folder, f"{os.path.splitext(os.path.basename(invoice_candidate))[0]}_invoice_ready.pdf")
                rotate_pdf_if_needed(invoice_candidate, invoice_final)
            elif ext in (".jpg", ".jpeg", ".png"):
                invoice_final = os.path.join(folder, f"{os.path.splitext(os.path.basename(invoice_candidate))[0]}_invoice_ready.pdf")
                image_to_pdf(invoice_candidate, invoice_final)
            elif ext in (".docx", ".doc"):
                invoice_final = os.path.join(folder, f"{os.path.splitext(os.path.basename(invoice_candidate))[0]}_invoice_ready.pdf")
                word_to_rotated_pdf(invoice_candidate, invoice_final)
            if logger:
                logger.log(f"Prepared invoice: {os.path.basename(invoice_final)}", "ok")
        except Exception as e:
            if logger:
                logger.log(f"Failed to prepare invoice {invoice_candidate}: {e}", "warn")
            invoice_final = None

    for p in other_files:
        root, ext = os.path.splitext(os.path.basename(p))
        ext = ext.lower()
        try:
            if ext == ".pdf":
                out_pdf = os.path.join(folder, f"{root}_ready.pdf")
                rotate_pdf_if_needed(p, out_pdf)
                prepared.append(out_pdf)
                if logger:
                    logger.log(f"Prepared PDF: {os.path.basename(p)}", "ok")
            elif ext in (".jpg", ".jpeg", ".png"):
                out_pdf = os.path.join(folder, f"{root}.pdf")
                image_to_pdf(p, out_pdf)
                prepared.append(out_pdf)
                if logger:
                    logger.log(f"Prepared Image: {os.path.basename(p)}", "ok")
            elif ext in (".docx", ".doc"):
                out_pdf = os.path.join(folder, f"{root}.pdf")
                word_to_rotated_pdf(p, out_pdf)
                prepared.append(out_pdf)
                if logger:
                    logger.log(f"Prepared Word: {os.path.basename(p)}", "ok")
        except Exception as e:
            if logger:
                logger.log(f"Error preparing {os.path.basename(p)}: {e}", "error")
            continue

    final_list = []
    if invoice_final:
        final_list.append(invoice_final)
    final_list.extend(prepared)
    return final_list, invoice_final


# ============ App ============
class App(tb.Window):
    def __init__(self):
        super().__init__(themename="flatly")
        self.title(APP_TITLE)
        self.geometry("1100x920")
        self.resizable(True, True)
        self.configure(bg=THEME_BG)

        outer = tb.Frame(self, bootstyle="secondary", padding=(10, 10))
        outer.pack(fill="both", expand=True, padx=10, pady=10)

        # Header
        header_frame = tk.Frame(outer, bg=PANEL_BG)
        header_frame.pack(fill="x", pady=(0, 8))
        header_label = tk.Label(header_frame, text=APP_TITLE, font=("Segoe UI", 20, "bold"),
                                bg=PANEL_BG, fg=TEXT_FG)
        header_label.pack(padx=8, pady=8, anchor="w")

        # TOP: titles and panels
        top_container = tk.Frame(outer, bg=THEME_BG)
        top_container.pack(fill="x", pady=(0, 10))

        lbl_clients = tk.Label(top_container, text="Clients", bg=THEME_BG, fg=TEXT_FG, font=("Segoe UI", 10, "bold"))
        lbl_clients.grid(row=0, column=0, sticky="w", padx=(2, 8))
        lbl_date = tk.Label(top_container, text="Week Ending (MM-DD)", bg=THEME_BG, fg=TEXT_FG, font=("Segoe UI", 10, "bold"))
        lbl_date.grid(row=0, column=1, sticky="w", padx=(8, 8))
        lbl_actions = tk.Label(top_container, text="Actions", bg=THEME_BG, fg=TEXT_FG, font=("Segoe UI", 10, "bold"))
        lbl_actions.grid(row=0, column=2, sticky="w", padx=(8, 8))

        # Clients panel
        clients_panel = tb.Frame(top_container, bootstyle="light", padding=8)
        clients_panel.grid(row=1, column=0, sticky="nsew", padx=(2, 8))
        top_container.grid_columnconfigure(0, weight=1)

        self.chk_vars = []
        for c in clients_list:
            var = tk.BooleanVar(value=True)
            cb = tb.Checkbutton(clients_panel, text=c, variable=var, bootstyle="round-toggle")
            try:
                cb.configure(foreground=TEXT_FG)
            except Exception:
                pass
            cb.pack(anchor="w", pady=2, padx=2)
            self.chk_vars.append((c, var))

        btns = tb.Frame(clients_panel)
        btns.pack(anchor="w", pady=(6, 0))
        tb.Button(btns, text="All", width=6, bootstyle="primary-outline",
                  command=lambda: [v.set(True) for _, v in self.chk_vars]).pack(side="left", padx=(0, 6))
        tb.Button(btns, text="None", width=6, bootstyle="secondary-outline",
                  command=lambda: [v.set(False) for _, v in self.chk_vars]).pack(side="left")

        # Date panel
        date_panel = tb.Frame(top_container, bootstyle="light", padding=8)
        date_panel.grid(row=1, column=1, sticky="nsew", padx=(8, 8))
        top_container.grid_columnconfigure(1, weight=1)

        mrow = tk.Frame(date_panel, bg=PANEL_BG)
        mrow.pack(anchor="w", pady=(2, 6))
        tk.Label(mrow, text="Month (MM):", bg=PANEL_BG, fg=TEXT_FG).pack(side="left")
        self.month_var = tk.StringVar(value=datetime.now().strftime("%m"))
        self.month_combo = ttk.Combobox(mrow, textvariable=self.month_var, width=8, state="normal")
        self.month_combo['values'] = [f"{i:02d}" for i in range(1, 13)]
        self.month_combo.pack(side="left", padx=(6, 12))

        drow = tk.Frame(date_panel, bg=PANEL_BG)
        drow.pack(anchor="w", pady=(2, 6))
        tk.Label(drow, text="Day (DD):", bg=PANEL_BG, fg=TEXT_FG).pack(side="left")
        self.day_var = tk.StringVar(value=datetime.now().strftime("%d"))
        self.day_combo = ttk.Combobox(drow, textvariable=self.day_var, width=8, state="normal")
        self.day_combo['values'] = [f"{i:02d}" for i in range(1, 32)]
        self.day_combo.pack(side="left", padx=(6, 12))

        tk.Label(date_panel, text="(searches for folders named 'Week MM-DD')", bg=PANEL_BG, fg=TEXT_FG).pack(anchor="w", pady=(6, 0))

        # Actions panel
        actions_panel = tb.Frame(top_container, bootstyle="light", padding=8)
        actions_panel.grid(row=1, column=2, sticky="nsew", padx=(8, 2))
        top_container.grid_columnconfigure(2, weight=0)

        self.start_btn = tb.Button(actions_panel, text="Start Merging", width=18, bootstyle="success", command=self.on_start)
        self.start_btn.pack(pady=(2, 6))
        tb.Button(actions_panel, text="Exit", width=18, bootstyle="danger", command=self.destroy).pack()

        # Middle area
        middle = tb.Frame(outer)
        middle.pack(fill="both", expand=True)

        left_col = tb.Frame(middle)
        left_col.pack(side="left", fill="both", expand=True, padx=(0, 8))

        prog_panel = tb.Labelframe(left_col, text="Progress", bootstyle="light", padding=10)
        prog_panel.pack(fill="x", pady=(0, 8))
        self.progress = tb.Progressbar(prog_panel, mode="determinate", length=700)
        self.progress.pack(fill="x", pady=(0, 6))
        try:
            self.progress.configure(maximum=100)
        except Exception:
            pass
        self.eta_label = tk.Label(prog_panel, text="0% • ETA: --", bg=PANEL_BG, fg=TEXT_FG)
        self.eta_label.pack(anchor="w")

        log_panel = tb.Labelframe(left_col, text="Log", bootstyle="light", padding=10)
        log_panel.pack(fill="both", expand=True)
        self.log_text = tk.Text(log_panel, height=18, wrap="word", bg=PANEL_BG, fg=TEXT_FG, relief="flat", font=("Segoe UI", 10))
        self.log_text.pack(fill="both", expand=True)
        self.log_text.tag_configure("ok", foreground="#1b8a5a")
        self.log_text.tag_configure("error", foreground="#c62828")
        self.log_text.tag_configure("warn", foreground="#b7791f")
        self.log_text.configure(state="disabled")

        right_col = tb.Labelframe(middle, text="Dashboard", bootstyle="light", padding=10)
        right_col.pack(side="right", fill="y", ipadx=8, ipady=8)

        tk.Label(right_col, text="Recent merges:", bg=PANEL_BG, fg=TEXT_FG).pack(anchor="w")
        self.recent_list = tk.Listbox(right_col, width=48, height=10, bg="white", fg=TEXT_FG)
        self.recent_list.pack(fill="x", pady=(4, 8))

        tb.Separator(right_col).pack(fill="x", pady=6)

        tk.Label(right_col, text="App Tips / Status:", bg=PANEL_BG, fg=TEXT_FG).pack(anchor="w")
        self.tips_text = tk.Text(right_col, height=8, bg=PANEL_BG, fg=TEXT_FG, relief="flat", wrap="word", font=("Segoe UI", 9))
        self.tips_text.pack(fill="both", expand=True)
        self.tips_text.insert("end", "- Use Refresh Scan when you change files.\n- Month and Day are editable; type MM and DD.\n- Start Merging will process each client and update Excel column G.\n")
        self.tips_text.configure(state="disabled")

        tb.Button(right_col, text="Refresh Scan", bootstyle="outline-primary", command=self.start_background_scan).pack(pady=(8, 0))

        self.pre_scan_info = {}
        self.total_tasks = 0
        self.tasks_done = 0
        self.scan_thread = None
        self._start_time = None

        # initial background scan
        self.start_background_scan()

    # Background scan
    def start_background_scan(self):
        if self.scan_thread and self.scan_thread.is_alive():
            return
        # schedule small UI note on main thread
        def note():
            self.log_text.configure(state="normal")
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.log_text.insert("end", f"[{ts}] Starting background scan...\n")
            self.log_text.see("end")
            self.log_text.configure(state="disabled")
        self.after(0, note)
        self.scan_thread = threading.Thread(target=self.refresh_scan, daemon=True)
        self.scan_thread.start()

    def refresh_scan(self):
        sel_clients = [name for name, v in self.chk_vars if v.get()]
        pre_scan = {}
        total_files = 0
        total_tasks = 0
        week_str = f"{self.month_combo.get()}-{self.day_combo.get()}"
        for c in sel_clients:
            client_root = os.path.join(main_folder, c)
            week_path = find_week_folder(client_root, week_str)
            if not week_path:
                pre_scan[c] = {"week": None, "files": [], "tasks": 1}
                total_tasks += 1
            else:
                raw_files = list_raw_files(week_path)
                tasks = 1 + len(raw_files) + 1 + 1
                pre_scan[c] = {"week": week_path, "files": raw_files, "tasks": tasks}
                total_files += len(raw_files)
                total_tasks += tasks

        def update_ui():
            self.pre_scan_info = pre_scan
            self.total_tasks = total_tasks
            header = f"Selected clients: {len(sel_clients)}    Total files: {total_files}    Estimated tasks: {total_tasks}\n"
            self.tips_text.configure(state="normal")
            try:
                # replace first line
                self.tips_text.delete("1.0", "2.0")
            except Exception:
                pass
            self.tips_text.insert("1.0", header)
            self.tips_text.configure(state="disabled")
            self.populate_recent_merges()
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.log_text.configure(state="normal")
            self.log_text.insert("end", f"[{ts}] Scan complete: {len(sel_clients)} clients, {total_files} files, {total_tasks} estimated tasks\n")
            self.log_text.see("end")
            self.log_text.configure(state="disabled")
        self.after(0, update_ui)

    def populate_recent_merges(self):
        # gather recent merged files safely on main thread
        def do_populate():
            recent = []
            try:
                for c in os.listdir(main_folder):
                    client_root = os.path.join(main_folder, c)
                    if not os.path.isdir(client_root):
                        continue
                    for m in os.listdir(client_root):
                        mpath = os.path.join(client_root, m)
                        if not os.path.isdir(mpath):
                            continue
                        for wk in os.listdir(mpath):
                            wkpath = os.path.join(mpath, wk)
                            if not os.path.isdir(wkpath):
                                continue
                            for name in os.listdir(wkpath):
                                if name.lower().endswith(".pdf") and name.endswith("_.pdf"):
                                    p = os.path.join(wkpath, name)
                                    try:
                                        t = os.path.getmtime(p)
                                        recent.append((t, p))
                                    except Exception:
                                        pass
            except Exception:
                pass
            recent.sort(reverse=True)
            self.recent_list.delete(0, "end")
            for _, p in recent[:10]:
                self.recent_list.insert("end", os.path.basename(p))
        # schedule on main thread
        self.after(0, do_populate)

    # Start merging
    def on_start(self):
        sel_clients = [name for name, v in self.chk_vars if v.get()]
        if not sel_clients:
            messagebox.showwarning("Select clients", "Please select at least one client.")
            return
        mm = self.month_combo.get().strip()
        dd = self.day_combo.get().strip()
        if not (mm.isdigit() and dd.isdigit() and 1 <= int(mm) <= 12 and 1 <= int(dd) <= 31):
            messagebox.showwarning("Select date", "Please enter a valid month and day (MM-DD).")
            return
        week_str = f"{mm.zfill(2)}-{dd.zfill(2)}"
        self.logger = StepLogger(self.log_text, week_str)

        if self.scan_thread and self.scan_thread.is_alive():
            self.scan_thread.join(timeout=1)

        self.total_tasks = sum(info.get("tasks", 0) for info in self.pre_scan_info.values())
        if self.total_tasks <= 0:
            self.logger.log("Nothing to do for selected clients and date.", "warn")
            return

        self.tasks_done = 0
        try:
            self.progress.configure(value=0)
        except Exception:
            pass
        self.eta_label.configure(text="0% • ETA: --")
        self.start_btn.configure(state="disabled")
        self._start_time = time.time()

        t = threading.Thread(target=self._run_merge, args=(list(self.pre_scan_info.items()), week_str), daemon=True)
        t.start()

    def _run_merge(self, pre_scan_items, week_str):
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            self.logger.log(f"Opened Excel: {excel_file}", "ok")
        except Exception as e:
            self.logger.log(f"Cannot open Excel: {e}", "error")
            self._finish(False)
            return

        for client, info in pre_scan_items:
            try:
                self.logger.log(f"--- Processing client: {client} ---")
                week_path = info.get("week")
                if not week_path:
                    self.logger.log(f"Week folder not found for {client} -> skipping", "warn")
                    self._increment_task_and_update()
                    continue
                else:
                    self.logger.log(f"Found week folder: {week_path}", "ok")
                    self._increment_task_and_update()

                try:
                    prepared_list, invoice_final = prepare_files_for_merge(week_path, logger=self.logger)
                except Exception as e:
                    self.logger.log(f"Error preparing files for {client}: {e}", "error")
                    prepared_list, invoice_final = [], None

                raw_files = info.get("files", [])
                for _ in raw_files:
                    self._increment_task_and_update()

                try:
                    if prepared_list:
                        merger = PdfMerger()
                        for p in prepared_list:
                            merger.append(p)
                        if invoice_final:
                            inv_base = os.path.splitext(os.path.basename(invoice_final))[0]
                            out_name = f"{inv_base}_.pdf"
                        else:
                            out_name = f"{client}_Week_{week_str}.pdf"
                        out_path = os.path.join(week_path, out_name)
                        merger.write(out_path)
                        merger.close()
                        self.logger.log(f"Merged PDF created: {out_path}", "ok")
                        # add to recent list safely
                        self.after(0, lambda p=out_path: self.recent_list.insert(0, os.path.basename(p)))
                    else:
                        out_path = ""
                        self.logger.log("No files to merge for this client.", "warn")
                except Exception as e:
                    out_path = ""
                    self.logger.log(f"Merge failed for {client}: {e}", "error")
                finally:
                    self._increment_task_and_update()

                try:
                    start_row, last_row = 4, ws.max_row
                    updated = False
                    for r in range(start_row, last_row + 1):
                        name_cell = ws.cell(row=r, column=2).value
                        if str(name_cell).strip() == client:
                            ws.cell(row=r, column=7).value = out_path
                            updated = True
                            break
                    if updated:
                        self.logger.log(f"Excel updated for {client} (Col G).", "ok")
                    else:
                        self.logger.log(f"Client '{client}' not found in Excel Col B (rows 4..{last_row}).", "warn")
                except Exception as e:
                    self.logger.log(f"Excel update failed for {client}: {e}", "error")
                finally:
                    self._increment_task_and_update()

            except Exception as e:
                self.logger.log(f"Client '{client}' failed unexpectedly: {e}", "error")
                continue

        try:
            wb.save(excel_file)
            self.logger.log("Excel saved.", "ok")
        except Exception as e:
            self.logger.log(f"Excel save error: {e}", "error")

        self._finish(True)

    def _increment_task_and_update(self):
        self.tasks_done += 1
        pct = int(self.tasks_done / max(1, self.total_tasks) * 100)
        elapsed = time.time() - getattr(self, "_start_time", time.time())
        avg = elapsed / max(1, self.tasks_done)
        remaining_seconds = int(avg * max(0, (self.total_tasks - self.tasks_done)))
        # schedule progress update on main thread
        self.progress.after(0, lambda v=pct: self.progress.configure(value=v))
        self.eta_label.after(0, lambda t=f"{pct}% • ETA: {remaining_seconds}s": self.eta_label.configure(text=t))

    def _finish(self, success: bool):
        try:
            winsound.MessageBeep(winsound.MB_OK)
        except Exception:
            try:
                winsound.Beep(750, 200)
            except Exception:
                pass
        msg = "Merging Completed ✅" if success else "Completed with Errors ⚠️"
        self.logger.log(msg, "ok" if success else "warn")
        self.start_btn.after(0, lambda: self.start_btn.configure(state="normal"))
        self.eta_label.after(0, lambda: self.eta_label.configure(text=("Done ✅" if success else "Done ⚠️")))


if __name__ == "__main__":
    app = App()
    app.title(APP_TITLE)
    app.configure(bg=THEME_BG)
    app.mainloop()
