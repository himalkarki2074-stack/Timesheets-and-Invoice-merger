import os
import sys
import time
import threading
from datetime import datetime
import tkinter as tk
from tkinter import messagebox
import ttkbootstrap as tb
from ttkbootstrap.constants import *

from PyPDF2 import PdfMerger, PdfReader, PdfWriter
from PIL import Image, ImageOps
from docx2pdf import convert
import openpyxl

# ===================== CONFIG =====================
main_folder = r"O:\ApTask\TDrive\FinTech LLC\Invoices\2025\Monthly"
excel_file  = r"C:\Users\HimalK\OneDrive - APTASK\Desktop\Aptask\Payroll\Automated email sheet\Emailexcel.xlsx"
log_folder  = r"C:\Users\HimalK\OneDrive - APTASK\Desktop\Aptask\Payroll\Timesheet_Invoice Merger\Logfile"
os.makedirs(log_folder, exist_ok=True)

clients_list = [
    "Aquila Energy", "BDR", "B Squared", "CFAIS", "Data Specialist",
    "HTS Workforce", "Schultz Controls", "Security 101", "VFS Fire", "Western Audio"
]

APP_TITLE = "Invoice and Timesheets Compiler"
TEXT_FG = "#000000"  # black text

# ===================== LOGGING =====================
class StepLogger:
    def __init__(self, text_widget, week_str):
        self.text_widget = text_widget
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.path = os.path.join(log_folder, f"Log_Week_{week_str}_{ts}.txt")

    def _write_file(self, line: str):
        with open(self.path, "a", encoding="utf-8") as f:
            f.write(line + "\n")

    def log(self, msg, tag="info"):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        line = f"[{timestamp}] {msg}"
        self.text_widget.configure(state="normal")
        if tag == "error":
            self.text_widget.insert("end", line + "\n", ("error",))
        elif tag == "ok":
            self.text_widget.insert("end", line + "\n", ("ok",))
        elif tag == "warn":
            self.text_widget.insert("end", line + "\n", ("warn",))
        else:
            self.text_widget.insert("end", line + "\n")
        self.text_widget.see("end")
        self.text_widget.configure(state="disabled")
        self._write_file(line)

# ===================== FILE HELPERS =====================
def rotate_pdf_if_needed(src_pdf, dst_pdf):
    reader = PdfReader(src_pdf)
    writer = PdfWriter()
    for page in reader.pages:
        try:
            w = float(page.mediabox.width)
            h = float(page.mediabox.height)
            if w > h:
                page.rotate(90)
        except:
            pass
        writer.add_page(page)
    with open(dst_pdf, "wb") as f:
        writer.write(f)

def image_to_pdf(image_path, out_pdf):
    image = Image.open(image_path)
    image = ImageOps.exif_transpose(image)
    image.convert("RGB").save(out_pdf)

def word_to_rotated_pdf(doc_path, out_pdf):
    temp_pdf = out_pdf.replace(".pdf", "_raw.pdf")
    convert(doc_path, temp_pdf)
    rotate_pdf_if_needed(temp_pdf, out_pdf)
    try:
        os.remove(temp_pdf)
    except:
        pass

def find_week_folder(client_root, week_str):
    target = f"Week {week_str}"
    if not os.path.isdir(client_root):
        return None
    for month_name in sorted(os.listdir(client_root)):
        month_path = os.path.join(client_root, month_name)
        if not os.path.isdir(month_path):
            continue
        week_path = os.path.join(month_path, target)
        if os.path.isdir(week_path):
            return week_path
    return None

def collect_files_for_merge(folder):
    invoice_pdf = None
    to_merge = []
    for name in sorted(os.listdir(folder)):
        path = os.path.join(folder, name)
        if not os.path.isfile(path):
            continue
        lower = name.lower()
        root, ext = os.path.splitext(name)
        ext = ext.lower()
        if "invoice" in lower and ext == ".pdf":
            invoice_pdf = path
            continue
        if ext in (".pdf", ".jpg", ".jpeg", ".png", ".docx", ".doc"):
            to_merge.append(path)

    prepared = []
    for p in to_merge:
        root, ext = os.path.splitext(os.path.basename(p))
        ext = ext.lower()
        try:
            if ext == ".pdf":
                rotated = os.path.join(folder, f"{root}_rotated.pdf")
                rotate_pdf_if_needed(p, rotated)
                prepared.append(rotated)
            elif ext in (".jpg", ".jpeg", ".png"):
                out_pdf = os.path.join(folder, f"{root}.pdf")
                image_to_pdf(p, out_pdf)
                prepared.append(out_pdf)
            elif ext in (".docx", ".doc"):
                out_pdf = os.path.join(folder, f"{root}.pdf")
                word_to_rotated_pdf(p, out_pdf)
                prepared.append(out_pdf)
        except:
            pass

    final_list = []
    if invoice_pdf:
        final_list.append(invoice_pdf)
    final_list.extend(prepared)
    return final_list, invoice_pdf

# ===================== APP =====================
class App(tb.Window):
    def __init__(self):
        super().__init__(themename="flatly")
        self.title(APP_TITLE)
        self.geometry("1020x720")
        self.configure(bg="#f0f0f0")

        outer = tb.Frame(self)
        outer.pack(fill="both", expand=True, padx=16, pady=16)

        # Header
        header = tb.Label(outer, text=APP_TITLE, font=("Segoe UI", 20, "bold"), bootstyle="light")
        header.pack(anchor="center", pady=(4,12), fill="x")

        # Top panel container
        top_row = tb.Frame(outer)
        top_row.pack(fill="x", pady=(0,10))

        # Clients Panel
        clients_panel = tb.Labelframe(top_row, text="Clients", bootstyle="secondary", padding=10)
        clients_panel.pack(side="left", fill="y", padx=(0,10))

        self.chk_vars = []
        for c in clients_list:
            var = tk.BooleanVar(value=True)
            cb = tb.Checkbutton(clients_panel, text=c, variable=var, bootstyle="round-toggle")
            cb.pack(anchor="w", pady=1)
            self.chk_vars.append((c, var))

        btns = tb.Frame(clients_panel)
        btns.pack(anchor="w", pady=(6,0))
        tb.Button(btns, text="All", width=6, bootstyle="primary-outline", command=lambda:[v.set(True) for _,v in self.chk_vars]).pack(side="left", padx=(0,6))
        tb.Button(btns, text="None", width=6, bootstyle="secondary-outline", command=lambda:[v.set(False) for _,v in self.chk_vars]).pack(side="left")

        # Week Panel
        date_panel = tb.Labelframe(top_row, text="Week Ending (MM-DD)", bootstyle="secondary", padding=10)
        date_panel.pack(side="left", fill="both", expand=True)

        mmrow = tb.Frame(date_panel)
        mmrow.pack(anchor="w", pady=2, fill="x")
        tb.Label(mmrow, text="Month (MM):", bootstyle="secondary").pack(side="left")
        self.month_var = tk.StringVar(value=datetime.now().strftime("%m"))
        self.month_entry = tb.Entry(mmrow, textvariable=self.month_var, width=6)
        self.month_entry.pack(side="left", padx=(6,0))

        ddrow = tb.Frame(date_panel)
        ddrow.pack(anchor="w", pady=2, fill="x")
        tb.Label(ddrow, text="Day (DD):", bootstyle="secondary").pack(side="left")
        self.day_var = tk.StringVar(value=datetime.now().strftime("%d"))
        self.day_entry = tb.Entry(ddrow, textvariable=self.day_var, width=6)
        self.day_entry.pack(side="left", padx=(34,0))

        # Actions Panel
        actions_panel = tb.Labelframe(top_row, text="Actions", bootstyle="secondary", padding=10)
        actions_panel.pack(side="right", fill="y")

        self.start_btn = tb.Button(actions_panel, text="Start Merging", width=16, bootstyle="success", command=self.on_start)
        self.start_btn.pack(pady=(2,6))
        tb.Button(actions_panel, text="Exit", width=16, bootstyle="danger", command=self.destroy).pack()

        # Middle Panel (Placeholder for stats / activity)
        middle_panel = tb.Labelframe(outer, text="Recent Activity / Stats", bootstyle="secondary", padding=10)
        middle_panel.pack(fill="both", expand=True, pady=(6,10))

        self.middle_text = tk.Text(middle_panel, height=10, wrap="word", bg="#ffffff", fg=TEXT_FG, relief="flat")
        self.middle_text.pack(fill="both", expand=True)
        self.middle_text.insert("end", "Recent tasks will appear here...\n")
        self.middle_text.configure(state="disabled")

        # Progress Panel
        prog_panel = tb.Labelframe(outer, text="Progress", bootstyle="secondary", padding=10)
        prog_panel.pack(fill="x", pady=(6,10))

        self.progress = tb.Progressbar(prog_panel, mode="determinate")
        self.progress.pack(fill="x")
        self.eta_label = tb.Label(prog_panel, text="0% • ETA: --", bootstyle="secondary")
        self.eta_label.pack(anchor="w", pady=(6,0))

        # Log Panel
        log_panel = tb.Labelframe(outer, text="Log", bootstyle="secondary", padding=10)
        log_panel.pack(fill="both", expand=True)

        self.log_text = tk.Text(log_panel, height=8, wrap="word", bg="#ffffff", fg=TEXT_FG, relief="flat")
        self.log_text.pack(fill="both", expand=True)
        self.log_text.tag_configure("ok", foreground="#1b8a5a")
        self.log_text.tag_configure("error", foreground="#c62828")
        self.log_text.tag_configure("warn", foreground="#b7791f")
        self.log_text.configure(state="disabled")

    def on_start(self):
        sel_clients = [name for name,v in self.chk_vars if v.get()]
        if not sel_clients:
            messagebox.showwarning("Select clients", "Please select at least one client.")
            return

        mm = self.month_var.get()
        dd = self.day_var.get()
        if not (mm and dd):
            messagebox.showwarning("Select date", "Please choose month and day (MM-DD).")
            return

        week_str = f"{mm}-{dd}"
        self.logger = StepLogger(self.log_text, week_str)
        self.start_btn.configure(state="disabled")
        self.progress.configure(value=0)
        self.eta_label.configure(text="0% • ETA: --")

        t = threading.Thread(target=self._run_merge, args=(sel_clients, week_str), daemon=True)
        t.start()

    def _run_merge(self, clients, week_str):
        start_time = time.time()
        total_tasks = len(clients)*3  # rough estimation: open folder, collect, merge
        task_done = 0

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            self.logger.log(f"Opened Excel: {excel_file}", "ok")
        except Exception as e:
            self.logger.log(f"Cannot open Excel: {e}", "error")
            self.start_btn.after(0, lambda: self.start_btn.configure(state="normal"))
            return

        for client in clients:
            try:
                self.logger.log(f"--- Processing client: {client} ---")
                client_root = os.path.join(main_folder, client)
                if not os.path.isdir(client_root):
                    self.logger.log(f"Client folder not found: {client_root}", "error")
                    raise FileNotFoundError(client_root)

                week_path = find_week_folder(client_root, week_str)
                if not week_path:
                    self.logger.log(f"Week folder not found under {client_root} -> 'Week {week_str}'", "warn")
                    raise FileNotFoundError(f"Week {week_str}")
                self.logger.log(f"Found week folder: {week_path}", "ok")
                task_done += 1
                self._update_progress(task_done, total_tasks)

                files_to_merge, invoice_pdf = collect_files_for_merge(week_path)
                if not files_to_merge:
                    self.logger.log("No valid files to merge.", "warn")
                    continue
                task_done += 1
                self._update_progress(task_done, total_tasks)

                merger = PdfMerger()
                for p in files_to_merge:
                    merger.append(p)
                if invoice_pdf:
                    inv_base = os.path.splitext(os.path.basename(invoice_pdf))[0]
                    out_name = f"{inv_base}_.pdf"
                else:
                    out_name = f"{client}_Week_{week_str}.pdf"
                out_path = os.path.join(week_path, out_name)
                merger.write(out_path)
                merger.close()
                self.logger.log(f"Merged PDF created: {out_path}", "ok")
                task_done += 1
                self._update_progress(task_done, total_tasks)

                # Update Excel
                try:
                    for r in range(4, ws.max_row+1):
                        if str(ws.cell(row=r, column=2).value).strip() == client:
                            ws.cell(row=r, column=7).value = out_path
                            break
                    self.logger.log(f"Excel updated for {client}.", "ok")
                except Exception as e:
                    self.logger.log(f"Excel update failed for {client}: {e}", "error")

            except Exception as e:
                self.logger.log(f"Client '{client}' failed: {e}", "error")

        try:
            wb.save(excel_file)
            self.logger.log("Excel saved.", "ok")
        except Exception as e:
            self.logger.log(f"Excel save error: {e}", "error")

        self.start_btn.after(0, lambda: self.start_btn.configure(state="normal"))
        self.eta_label.after(0, lambda: self.eta_label.configure(text="Done ✅"))
        self.progress.after(0, lambda: self.progress.configure(value=100))

    def _update_progress(self, done, total):
        pct = int(done/total*100)
        self.progress.after(0, lambda: self.progress.configure(value=pct))
        self.eta_label.after(0, lambda: self.eta_label.configure(text=f"{pct}% • ETA: --"))

if __name__ == "__main__":
    app = App()
    app.mainloop()
