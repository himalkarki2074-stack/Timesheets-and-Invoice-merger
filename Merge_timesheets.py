import os
from PyPDF2 import PdfMerger, PdfReader, PdfWriter
from PIL import Image, ImageOps
from docx2pdf import convert
from tqdm import tqdm
import openpyxl

# ====== CONFIGURATION ======
main_folder = r"O:\ApTask\TDrive\FinTech LLC\Invoices\2025\Monthly"
log_folder = r"C:\Users\HimalK\OneDrive - APTASK\Desktop\Aptask\Payroll\Timesheet_Invoice Merger\Logfile"
excel_file = r"C:\Users\HimalK\OneDrive - APTASK\Desktop\Aptask\Payroll\Automated email sheet\Emailexcel.xlsx"
os.makedirs(log_folder, exist_ok=True)

# ====== USER INPUT ======
clients_input = input("Enter client folder names (comma separated): ")
clients_to_process = [c.strip() for c in clients_input.split(",")]
week_to_process = input("Enter the week folder to process (e.g., Week 08-03): ").strip()

# ====== LOG FILE ======
log_file_path = os.path.join(log_folder, f"Logfile_{week_to_process.replace(' ', '_')}.txt")
log_entries = []

# ====== OPEN EXCEL ======
wb = openpyxl.load_workbook(excel_file)
ws = wb.active
start_row = 4
last_row = ws.max_row

# ====== LOOP THROUGH CLIENTS ======
for row in range(start_row, last_row + 1):
    client_name = ws.cell(row=row, column=2).value  # Column B has client names
    if not client_name or client_name not in clients_to_process:
        continue

    client_path = os.path.join(main_folder, client_name)
    if not os.path.isdir(client_path):
        print(f"Client folder not found: {client_name}")
        log_entries.append(f"Client folder not found: {client_name}")
        continue

    merged_invoice_path = None

    # LOOP THROUGH MONTHS
    for month in os.listdir(client_path):
        month_path = os.path.join(client_path, month)
        if not os.path.isdir(month_path):
            continue

        week_path = os.path.join(month_path, week_to_process)
        if os.path.isdir(week_path):
            folder_path = week_path
            print(f"\nProcessing folder: {folder_path}")
            log_entries.append(f"\nProcessing folder: {folder_path}")

            files = os.listdir(folder_path)
            invoice_file = None
            timesheet_files = []

            for f in files:
                if f.lower().endswith('.pdf') and "invoice" in f.lower():
                    invoice_file = f
                elif f.lower().endswith(('.pdf', '.jpg', '.jpeg', '.png', '.docx', '.doc')):
                    timesheet_files.append(f)

            if not invoice_file:
                print(f"Invoice not found in {folder_path}")
                log_entries.append(f"Invoice not found in {folder_path}")
                continue

            # ====== CONVERT TIMESHEETS TO PDF AND ROTATE ======
            pdf_files_to_merge = []
            pdf_files_to_merge.append(os.path.join(folder_path, invoice_file))

            print("Converting and rotating timesheets if needed...")
            log_entries.append("Converting and rotating timesheets if needed...")

            for ts in tqdm(timesheet_files, desc=f"Processing timesheets for {client_name}", unit="file"):
                ts_path = os.path.join(folder_path, ts)
                filename, ext = os.path.splitext(ts)
                ext = ext.lower()

                try:
                    if ext == '.pdf':
                        reader = PdfReader(ts_path)
                        writer = PdfWriter()
                        for page in reader.pages:
                            if page.mediabox.width > page.mediabox.height:
                                page.rotate(90)
                            writer.add_page(page)
                        rotated_pdf_path = os.path.join(folder_path, f"{filename}_rotated.pdf")
                        with open(rotated_pdf_path, "wb") as f:
                            writer.write(f)
                        pdf_files_to_merge.append(rotated_pdf_path)
                        log_entries.append(f"Processed PDF: {ts}")

                    elif ext in ['.jpg', '.jpeg', '.png']:
                        image = Image.open(ts_path)
                        image = ImageOps.exif_transpose(image)
                        if image.width > image.height:
                            image = image.rotate(90, expand=True)
                        pdf_path = os.path.join(folder_path, f"{filename}.pdf")
                        image.convert('RGB').save(pdf_path)
                        pdf_files_to_merge.append(pdf_path)
                        log_entries.append(f"Processed Image: {ts}")

                    elif ext in ['.docx', '.doc']:
                        pdf_path = os.path.join(folder_path, f"{filename}.pdf")
                        convert(ts_path, pdf_path)
                        reader = PdfReader(pdf_path)
                        writer = PdfWriter()
                        for page in reader.pages:
                            if page.mediabox.width > page.mediabox.height:
                                page.rotate(90)
                            writer.add_page(page)
                        rotated_pdf_path = os.path.join(folder_path, f"{filename}_rotated.pdf")
                        with open(rotated_pdf_path, "wb") as f:
                            writer.write(f)
                        pdf_files_to_merge.append(rotated_pdf_path)
                        log_entries.append(f"Processed Word file: {ts}")

                except Exception as e:
                    print(f"Error processing {ts}: {str(e)}")
                    log_entries.append(f"Error processing {ts}: {str(e)}")

            # ====== MERGE PDFs ======
            merger = PdfMerger()
            for pdf in pdf_files_to_merge:
                merger.append(pdf)

            invoice_name = os.path.splitext(invoice_file)[0]
            merged_invoice_path = os.path.join(folder_path, f"{invoice_name}_.pdf")
            merger.write(merged_invoice_path)
            merger.close()
            print(f"Merged PDF created: {merged_invoice_path}")
            log_entries.append(f"Merged PDF created: {merged_invoice_path}")

        else:
            print(f"Week folder not found: {week_to_process} in {month}")
            log_entries.append(f"Week folder not found: {week_to_process} in {month}")

    # ====== UPDATE EXCEL ======
    if merged_invoice_path:
        ws.cell(row=row, column=7).value = merged_invoice_path
        log_entries.append(f"Excel updated for client {client_name}: {merged_invoice_path}")
    else:
        log_entries.append(f"No merged invoice to update Excel for client {client_name}")

# ====== SAVE EXCEL ======
wb.save(excel_file)
log_entries.append("\nProcess completed successfully.")

# ====== SAVE LOG FILE ======
with open(log_file_path, "w", encoding="utf-8") as log_file:
    log_file.write("\n".join(log_entries))

print(f"\nLog file saved: {log_file_path}")
