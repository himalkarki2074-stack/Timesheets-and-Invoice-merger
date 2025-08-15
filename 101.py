import os
import sys
import time
import threading
import winsound
from datetime import datetime

# --- UI (Windows look) ---
import tkinter as tk
from tkinter import messagebox
import ttkbootstrap as tb
from ttkbootstrap.constants import *

# --- File handling / conversions ---
from PyPDF2 import PdfMerger, PdfReader, PdfWriter
from PIL import Image, ImageOps
from docx2pdf import convert
import openpyxl

# ===================== CONFIG =====================
# Folder layout assumed:
# main_folder / <Client> / <Month folder> / Week MM-DD / (Invoice + Timesheets)
main_folder = r"O:\ApTask\TDrive\FinTech LLC\Invoices\2025\Monthly"
excel_file  = r"C:\Users\HimalK\OneDrive - APTASK\Desktop\Aptask\Payroll\Automated email sheet\Emailexcel.xlsx"
log_folder  = r"C:\Users\HimalK\OneDrive - APTASK\Desktop\Aptask\Payroll\Timesheet_Invoice Merger\Logfile"
os.makedirs(log_folder, exist_ok=True)

clients_list = [
    "Aquila Energy", "BDR", "B Squared", "CFAIS", "Data Specialist",
    "HTS Workforce", "Schultz Controls", "Security 101", "VFS Fire", "Western Audio"
]

APP_TITLE = "Invoice Timesheets Merger"  # window + header text
THEME_BG  = "#2b6cb0"   # medium blue background
PANEL_BG  = "#ffffff"   # white panels for high contrast
TEXT_FG   = "#0b2240"   # dark blue text
ACCENT    = "#1e90ff"   # accent blue
# ===================================================


# ------------- Utility: Logging to UI + file -------------
class StepLogger:
    def __init__(self, text_widget, week_str):
        self.text_widget = text_widget
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.path = os.path.join(log_folder, f"Log_Week_{week_str}_{ts}.txt")

    def _write_file(self, line: str):
        with open(self.path, "a", encoding="utf-8") as f:
            f.write(line + "\n")

    def log(self, msg, tag="info"):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        line = f"[{timestamp}] {msg}"
        # UI
        self.text_widget.configure(state="normal")
        if tag == "error":
            self.text_widget.insert("end", line + "\n", ("error",))
        elif tag == "ok":
            self.text_widget.insert("end", line + "\n", ("ok",))
        elif tag == "warn":
            self.text_widget.insert("end", line + "\n", ("warn",))
        else:
            self.text_widget.insert("end", line + "\n")
        self.text_widget.see("end")
        self.text_widget.configure(state="disabled")
        # File
        self._write_file(line)


# ------------- Core merging helpers -------------
def rotate_pdf_if_needed(src_pdf, dst_pdf):
    """Rotate landscape pages to portrait if needed."""
    reader = PdfReader(src_pdf)
    writer = PdfWriter()
    for page in reader.pages:
        try:
            w = float(page.mediabox.width)
            h = float(page.mediabox.height)
            if w > h:
                page.rotate(90)
        except Exception:
            pass
        writer.add_page(page)
    with open(dst_pdf, "wb") as f:
        writer.write(f)


def image_to_pdf(image_path, out_pdf):
    """Normalize orientation and save as PDF."""
    image = Image.open(image_path)
    image = ImageOps.exif_transpose(image)
    image.convert("RGB").save(out_pdf)


def word_to_rotated_pdf(doc_path, out_pdf):
    """Convert Word to PDF via Word (Windows), then rotate pages if needed."""
    temp_pdf = out_pdf.replace(".pdf", "_raw.pdf")
    convert(doc_path, temp_pdf)  # requires Word installed
    rotate_pdf_if_needed(temp_pdf, out_pdf)
    try:
        os.remove(temp_pdf)
    except Exception:
        pass


def find_week_folder(client_root, week_str):
    """
    Folder structure: client_root/<Month folder>/Week MM-DD
    We scan month folders and pick the first match for 'Week {week_str}'.
    """
    target = f"Week {week_str}"
    if not os.path.isdir(client_root):
        return None
    for month_name in sorted(os.listdir(client_root)):
        month_path = os.path.join(client_root, month_name)
        if not os.path.isdir(month_path):
            continue
        week_path = os.path.join(month_path, target)
        if os.path.isdir(week_path):
            return week_path
    return None


def collect_files_for_merge(folder):
    """
    Identify the invoice (prefer 'invoice' in name, .pdf) and any timesheet sources.
    Convert/rotate as needed, return ordered list with invoice first if found.
    """
    invoice_pdf = None
    to_merge = []

    for name in sorted(os.listdir(folder)):
        path = os.path.join(folder, name)
        if not os.path.isfile(path):
            continue
        lower = name.lower()
        root, ext = os.path.splitext(name)
        ext = ext.lower()

        # Prefer invoice pdf first
        if "invoice" in lower and ext == ".pdf":
            invoice_pdf = path
            continue

        # Other acceptable sources
        if ext in (".pdf", ".jpg", ".jpeg", ".png", ".docx", ".doc"):
            to_merge.append(path)

    # Now convert all non-PDFs and normalize PDFs
    prepared = []
    for p in to_merge:
        root, ext = os.path.splitext(os.path.basename(p))
        ext = ext.lower()

        try:
            if ext == ".pdf":
                rotated = os.path.join(folder, f"{root}_rotated.pdf")
                rotate_pdf_if_needed(p, rotated)
                prepared.append(rotated)
            elif ext in (".jpg", ".jpeg", ".png"):
                out_pdf = os.path.join(folder, f"{root}.pdf")
                image_to_pdf(p, out_pdf)
                prepared.append(out_pdf)
            elif ext in (".docx", ".doc"):
                out_pdf = os.path.join(folder, f"{root}.pdf")
                word_to_rotated_pdf(p, out_pdf)
                prepared.append(out_pdf)
        except Exception as e:
            # skip problematic file
            pass

    final_list = []
    if invoice_pdf:
        final_list.append(invoice_pdf)
    final_list.extend(prepared)
    return final_list, invoice_pdf


# ------------- UI App -------------
class App(tb.Window):
    def __init__(self):
        super().__init__(themename="flatly")  # Windows-like, clean
        self.title(APP_TITLE)
        self.geometry("1020x700")
        self.configure(bg=THEME_BG)

        # Main container
        outer = tb.Frame(self, bootstyle="secondary")
        outer.pack(fill="both", expand=True, padx=16, pady=16)

        # Header
        header = tb.Label(
            outer, text=APP_TITLE,
            font=("Segoe UI", 20, "bold"),
            bootstyle="light"
        )
        header.pack(anchor="center", pady=(4, 12))

        # Top controls row
        top_row = tb.Frame(outer)
        top_row.pack(fill="x", pady=(0, 10))

        # Clients panel
        clients_panel = tb.Labelframe(
            top_row, text="Clients", bootstyle="info",
            padding=10
        )
        clients_panel.pack(side="left", fill="y", padx=(0, 10))

        # Checkbox list
        self.chk_vars = []
        for c in clients_list:
            var = tk.BooleanVar(value=True)
            cb = tb.Checkbutton(
                clients_panel, text=c, variable=var,
                bootstyle="round-toggle"  # visible check style
            )
            cb.pack(anchor="w", pady=1)
            self.chk_vars.append((c, var))

        # Small select/deselect buttons
        btns = tb.Frame(clients_panel)
        btns.pack(anchor="w", pady=(6, 0))
        tb.Button(btns, text="All", width=6, bootstyle="primary-outline",
                  command=lambda: [v.set(True) for _, v in self.chk_vars]).pack(side="left", padx=(0, 6))
        tb.Button(btns, text="None", width=6, bootstyle="secondary-outline",
                  command=lambda: [v.set(False) for _, v in self.chk_vars]).pack(side="left")

        # Date panel (US format MM-DD)
        date_panel = tb.Labelframe(
            top_row, text="Week Ending (US format)", bootstyle="info",
            padding=10
        )
        date_panel.pack(side="left", fill="both", expand=True)

        # Month dropdown 01..12
        mrow = tb.Frame(date_panel)
        mrow.pack(anchor="w", pady=2, fill="x")
        tb.Label(mrow, text="Month (MM):", bootstyle="light").pack(side="left")
        self.month_var = tk.StringVar(value=datetime.now().strftime("%m"))
        self.month_combo = tb.Combobox(mrow, textvariable=self.month_var, width=6, state="readonly")
        self.month_combo["values"] = [f"{i:02d}" for i in range(1, 13)]
        self.month_combo.pack(side="left", padx=(6, 0))

        # Day dropdown 01..31
        drow = tb.Frame(date_panel)
        drow.pack(anchor="w", pady=2, fill="x")
        tb.Label(drow, text="Day (DD):", bootstyle="light").pack(side="left")
        self.day_var = tk.StringVar(value=datetime.now().strftime("%d"))
        self.day_combo = tb.Combobox(drow, textvariable=self.day_var, width=6, state="readonly")
        self.day_combo["values"] = [f"{i:02d}" for i in range(1, 32)]
        self.day_combo.pack(side="left", padx=(34, 0))

        # Hint
        tb.Label(date_panel, text="This finds folders named: Week MM-DD",
                 bootstyle="secondary").pack(anchor="w", pady=(8, 0))

        # Actions panel
        actions = tb.Labelframe(top_row, text="Actions", bootstyle="info", padding=10)
        actions.pack(side="right", fill="y")

        self.start_btn = tb.Button(actions, text="Start Merging", width=16,
                                   bootstyle="success", command=self.on_start)
        self.start_btn.pack(pady=(2, 6))

        tb.Button(actions, text="Exit", width=16, bootstyle="danger",
                  command=self.destroy).pack()

        # Progress + ETA
        prog_panel = tb.Labelframe(outer, text="Progress", bootstyle="info", padding=10)
        prog_panel.pack(fill="x", pady=(6, 10))

        self.progress = tb.Progressbar(prog_panel, mode="determinate")
        self.progress.pack(fill="x")
        self.eta_label = tb.Label(prog_panel, text="0% • ETA: --", bootstyle="secondary")
        self.eta_label.pack(anchor="w", pady=(6, 0))

        # Log panel
        log_panel = tb.Labelframe(outer, text="Log", bootstyle="info", padding=10)
        log_panel.pack(fill="both", expand=True)

        self.log_text = tk.Text(
            log_panel, height=16, wrap="word",
            bg=PANEL_BG, fg=TEXT_FG, relief="flat"
        )
        self.log_text.pack(fill="both", expand=True)
        self.log_text.tag_configure("ok", foreground="#1b8a5a")
        self.log_text.tag_configure("error", foreground="#c62828")
        self.log_text.tag_configure("warn", foreground="#b7791f")
        self.log_text.configure(state="disabled")

        # Footer note
        tb.Label(outer, text=f"Root: {main_folder}", bootstyle="light").pack(anchor="w")

    # ------------ UI handlers ------------
    def on_start(self):
        sel_clients = [name for name, v in self.chk_vars if v.get()]
        if not sel_clients:
            messagebox.showwarning("Select clients", "Please select at least one client.")
            return

        mm = self.month_var.get()
        dd = self.day_var.get()
        if not (mm and dd):
            messagebox.showwarning("Select date", "Please choose month and day (MM-DD).")
            return

        week_str = f"{mm}-{dd}"  # US format for folder: "Week MM-DD"
        self.logger = StepLogger(self.log_text, week_str)

        self.start_btn.configure(state="disabled")
        self.progress.configure(value=0)
        self.eta_label.configure(text="0% • ETA: --")

        t = threading.Thread(target=self._run_merge, args=(sel_clients, week_str), daemon=True)
        t.start()

    # ------------ Merge logic (thread) ------------
    def _run_merge(self, clients, week_str):
        start_time = time.time()
        total = len(clients)
        done = 0

        # Open Excel
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            self.logger.log(f"Opened Excel: {excel_file}", "ok")
        except Exception as e:
            self.logger.log(f"Cannot open Excel: {e}", "error")
            self._finish(False)
            return

        # Process each client
        for client in clients:
            try:
                self.logger.log(f"--- Processing client: {client} ---")
                client_root = os.path.join(main_folder, client)
                if not os.path.isdir(client_root):
                    self.logger.log(f"Client folder not found: {client_root}", "error")
                    raise FileNotFoundError(client_root)

                week_path = find_week_folder(client_root, week_str)
                if not week_path:
                    self.logger.log(f"Week folder not found under {client_root} -> 'Week {week_str}'", "warn")
                    raise FileNotFoundError(f"Week {week_str}")

                self.logger.log(f"Found week folder: {week_path}", "ok")

                files_to_merge, invoice_pdf = collect_files_for_merge(week_path)
                if not files_to_merge:
                    self.logger.log("No valid files to merge (PDF/Image/Word).", "warn")
                    raise RuntimeError("Nothing to merge")

                # Merge now
                merger = PdfMerger()
                for p in files_to_merge:
                    merger.append(p)

                # Output name (put invoice base name if found, else generic)
                if invoice_pdf:
                    inv_base = os.path.splitext(os.path.basename(invoice_pdf))[0]
                    out_name = f"{inv_base}_.pdf"
                else:
                    out_name = f"{client}_Week_{week_str}.pdf"
                out_path = os.path.join(week_path, out_name)

                merger.write(out_path)
                merger.close()
                self.logger.log(f"Merged PDF created: {out_path}", "ok")

                # Update Excel column G with the merged path
                # Excel: find client name in Column B (row 4..max)
                try:
                    start_row, last_row = 4, ws.max_row
                    updated = False
                    for r in range(start_row, last_row + 1):
                        name_cell = ws.cell(row=r, column=2).value  # Column B
                        if str(name_cell).strip() == client:
                            ws.cell(row=r, column=7).value = out_path  # Column G
                            updated = True
                            break
                    if updated:
                        self.logger.log(f"Excel updated for {client} (Col G): {out_path}", "ok")
                    else:
                        self.logger.log(f"Client '{client}' not found in Excel Col B (rows 4..{last_row}).", "warn")
                except Exception as e:
                    self.logger.log(f"Excel update failed for {client}: {e}", "error")

            except Exception as e:
                self.logger.log(f"Client '{client}' failed: {e}", "error")

            # Progress & ETA
            done += 1
            elapsed = max(0.001, time.time() - start_time)
            avg = elapsed / done
            remaining = int(avg * (total - done))
            pct = int(done / total * 100)

            self.progress.after(0, lambda v=pct: self.progress.configure(value=v))
            self.eta_label.after(0, lambda t=f"{pct}% • ETA: {remaining}s": self.eta_label.configure(text=t))

        # Save Excel once after loop
        try:
            wb.save(excel_file)
            self.logger.log("Excel saved.", "ok")
        except Exception as e:
            self.logger.log(f"Excel save error: {e}", "error")

        # Done!
        self._finish(True)

    def _finish(self, success: bool):
        # 1-second notification beep (standard)
        try:
            winsound.Beep(1000, 1000)
        except Exception:
            pass

        msg = "Merging Completed ✅" if success else "Completed with Errors ⚠️"
        self.logger.log(msg, "ok" if success else "warn")
        self.start_btn.after(0, lambda: self.start_btn.configure(state="normal"))
        self.eta_label.after(0, lambda: self.eta_label.configure(text=("Done ✅" if success else "Done ⚠️")))


# ----------------- run app -----------------
if __name__ == "__main__":
    app = App()
    app.title(APP_TITLE)  # ensure Windows-style chrome (not mac/iOS)
    # Force medium blue background on root window
    app.configure(bg=THEME_BG)
    app.mainloop()
