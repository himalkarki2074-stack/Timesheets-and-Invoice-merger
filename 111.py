# invoice_timesheets_compiler_v110.py
import os
import time
import threading
from datetime import datetime
import tkinter as tk
from tkinter import messagebox
import ttkbootstrap as tb
from ttkbootstrap.constants import *

from PyPDF2 import PdfMerger, PdfReader, PdfWriter
from PIL import Image, ImageOps
from docx2pdf import convert
import openpyxl
import winsound

# ================== CONFIG ==================
main_folder = r"O:\ApTask\TDrive\FinTech LLC\Invoices\2025\Monthly"
excel_file  = r"C:\Users\HimalK\OneDrive - APTASK\Desktop\Aptask\Payroll\Automated email sheet\Emailexcel.xlsx"
log_folder  = r"C:\Users\HimalK\OneDrive - APTASK\Desktop\Aptask\Payroll\Timesheet_Invoice Merger\Logfile"
os.makedirs(log_folder, exist_ok=True)

clients_list = [
    "Aquila Energy", "BDR", "B Squared", "CFAIS", "Data Specialist",
    "HTS Workforce", "Schultz Controls", "Security 101", "VFS Fire", "Western Audio"
]

APP_TITLE = "Invoice and Timesheets Compiler"
THEME_BG = "#2b6cb0"     # medium blue window background (used sparingly)
PANEL_BG = "#ffffff"     # white panels
TEXT_FG  = "#000000"     # black text everywhere except buttons
# ============================================


# ---------- Thread-safe logger (schedules UI updates) ----------
class StepLogger:
    def __init__(self, text_widget, week_str):
        self.text_widget = text_widget
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_week = week_str.replace("/", "-")
        self.path = os.path.join(log_folder, f"Log_Week_{safe_week}_{ts}.txt")
        # ensure file exists
        try:
            with open(self.path, "a", encoding="utf-8"):
                pass
        except Exception:
            pass

    def _write_file(self, line: str):
        try:
            with open(self.path, "a", encoding="utf-8") as f:
                f.write(line + "\n")
        except Exception:
            pass

    def log(self, msg, tag="info"):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        line = f"[{timestamp}] {msg}"
        # schedule UI update on main thread
        def ui_append():
            try:
                self.text_widget.configure(state="normal")
                if tag in ("error", "ok", "warn"):
                    self.text_widget.insert("end", line + "\n", (tag,))
                else:
                    self.text_widget.insert("end", line + "\n")
                self.text_widget.see("end")
                self.text_widget.configure(state="disabled")
            except Exception:
                pass
        try:
            self.text_widget.after(0, ui_append)
        except Exception:
            ui_append()
        # write to file
        self._write_file(line)


# ---------- File helpers ----------
def rotate_pdf_if_needed(src_pdf, dst_pdf):
    try:
        reader = PdfReader(src_pdf)
        writer = PdfWriter()
        for page in reader.pages:
            try:
                w = float(page.mediabox.width)
                h = float(page.mediabox.height)
                if w > h:
                    page.rotate(90)
            except Exception:
                pass
            writer.add_page(page)
        with open(dst_pdf, "wb") as f:
            writer.write(f)
    except Exception:
        # best-effort: copy original if rotation fails
        try:
            from shutil import copyfile
            copyfile(src_pdf, dst_pdf)
        except Exception:
            pass


def image_to_pdf(image_path, out_pdf):
    image = Image.open(image_path)
    image = ImageOps.exif_transpose(image)
    image.convert("RGB").save(out_pdf)


def word_to_rotated_pdf(doc_path, out_pdf):
    temp_pdf = out_pdf.replace(".pdf", "_raw.pdf")
    convert(doc_path, temp_pdf)  # requires MS Word on Windows
    rotate_pdf_if_needed(temp_pdf, out_pdf)
    try:
        os.remove(temp_pdf)
    except Exception:
        pass


def find_week_folder(client_root, week_str):
    if not os.path.isdir(client_root):
        return None
    target = f"Week {week_str}"
    for month_name in sorted(os.listdir(client_root)):
        month_path = os.path.join(client_root, month_name)
        if not os.path.isdir(month_path):
            continue
        week_path = os.path.join(month_path, target)
        if os.path.isdir(week_path):
            return week_path
    return None


def list_raw_files(folder):
    files = []
    if not os.path.isdir(folder):
        return files
    for name in sorted(os.listdir(folder)):
        path = os.path.join(folder, name)
        if not os.path.isfile(path):
            continue
        ext = os.path.splitext(name)[1].lower()
        if ext in (".pdf", ".jpg", ".jpeg", ".png", ".docx", ".doc"):
            files.append(path)
    return files


def prepare_files_for_merge(folder, logger=None):
    """
    Convert/rotate files and prepare list. Also ensures invoice (any file with 'invoice' in name)
    is converted if needed and placed first in the returned list.
    """
    invoice_candidate = None
    others = []
    for name in sorted(os.listdir(folder)):
        path = os.path.join(folder, name)
        if not os.path.isfile(path):
            continue
        lower = name.lower()
        ext = os.path.splitext(name)[1].lower()
        if "invoice" in lower:
            invoice_candidate = path
            continue
        if ext in (".pdf", ".jpg", ".jpeg", ".png", ".docx", ".doc"):
            others.append(path)

    prepared = []
    invoice_final = None

    # prepare invoice (if found)
    if invoice_candidate:
        root = os.path.splitext(os.path.basename(invoice_candidate))[0]
        invoice_final = os.path.join(folder, f"{root}_invoice_ready.pdf")
        try:
            if invoice_candidate.lower().endswith(".pdf"):
                rotate_pdf_if_needed(invoice_candidate, invoice_final)
            elif invoice_candidate.lower().endswith((".jpg", ".jpeg", ".png")):
                image_to_pdf(invoice_candidate, invoice_final)
            elif invoice_candidate.lower().endswith((".docx", ".doc")):
                word_to_rotated_pdf(invoice_candidate, invoice_final)
            if logger:
                logger.log(f"Prepared invoice: {os.path.basename(invoice_final)}", "ok")
        except Exception as e:
            invoice_final = None
            if logger:
                logger.log(f"Failed to prepare invoice {invoice_candidate}: {e}", "warn")

    # prepare other files
    for p in others:
        root = os.path.splitext(os.path.basename(p))[0]
        ext = os.path.splitext(p)[1].lower()
        try:
            if ext == ".pdf":
                out_pdf = os.path.join(folder, f"{root}_rotated.pdf")
                rotate_pdf_if_needed(p, out_pdf)
                prepared.append(out_pdf)
                if logger:
                    logger.log(f"Prepared PDF: {os.path.basename(p)}", "ok")
            elif ext in (".jpg", ".jpeg", ".png"):
                out_pdf = os.path.join(folder, f"{root}.pdf")
                image_to_pdf(p, out_pdf)
                prepared.append(out_pdf)
                if logger:
                    logger.log(f"Prepared Image: {os.path.basename(p)}", "ok")
            elif ext in (".docx", ".doc"):
                out_pdf = os.path.join(folder, f"{root}.pdf")
                word_to_rotated_pdf(p, out_pdf)
                prepared.append(out_pdf)
                if logger:
                    logger.log(f"Prepared Word: {os.path.basename(p)}", "ok")
        except Exception as e:
            if logger:
                logger.log(f"Error preparing {os.path.basename(p)}: {e}", "error")
            continue

    final_list = []
    if invoice_final:
        final_list.append(invoice_final)
    final_list.extend(prepared)
    return final_list, invoice_final


# ============ APPLICATION ============
class App(tb.Window):
    def __init__(self):
        super().__init__(themename="flatly")
        self.title(APP_TITLE)
        # prefer a tall window that fits content well
        self.geometry("1120x820")
        self.configure(bg=THEME_BG)

        # root frame (ttkbootstrap frame)
        root = tb.Frame(self, padding=12)
        root.pack(fill="both", expand=True)

        # Header (white stripe so black title is visible)
        header_frame = tk.Frame(root, bg=PANEL_BG)
        header_frame.pack(fill="x", pady=(0, 10))
        header_label = tk.Label(header_frame, text=APP_TITLE, font=("Segoe UI", 20, "bold"),
                                bg=PANEL_BG, fg=TEXT_FG, anchor="w")
        header_label.pack(fill="x", padx=8, pady=10)

        # Top row: titles (explicit top labels) + panels under them
        top_container = tk.Frame(root, bg=THEME_BG)
        top_container.pack(fill="x", pady=(0, 6))

        # Top labels
        lbl_clients = tk.Label(top_container, text="Clients", bg=THEME_BG, fg=TEXT_FG, font=("Segoe UI", 10, "bold"))
        lbl_clients.grid(row=0, column=0, sticky="w", padx=(4, 8))
        lbl_week = tk.Label(top_container, text="Week Ending (MM-DD)", bg=THEME_BG, fg=TEXT_FG, font=("Segoe UI", 10, "bold"))
        lbl_week.grid(row=0, column=1, sticky="w", padx=(8, 8))
        lbl_actions = tk.Label(top_container, text="Actions", bg=THEME_BG, fg=TEXT_FG, font=("Segoe UI", 10, "bold"))
        lbl_actions.grid(row=0, column=2, sticky="w", padx=(8, 8))

        # Panels row
        # Clients panel (left)
        clients_panel = tb.Frame(top_container, bootstyle="light", padding=8)
        clients_panel.grid(row=1, column=0, sticky="nsew", padx=(4, 8))
        top_container.grid_columnconfigure(0, weight=1)

        # client checkboxes
        self.chk_vars = []
        for c in clients_list:
            var = tk.BooleanVar(value=True)
            cb = tb.Checkbutton(clients_panel, text=c, variable=var, bootstyle="round-toggle")
            # many ttk widgets ignore fg param; using default label color is fine since header is black
            cb.pack(anchor="w", pady=2)
            self.chk_vars.append((c, var))

        # small select/deselect
        sel_btns = tb.Frame(clients_panel)
        sel_btns.pack(anchor="w", pady=(6, 0))
        tb.Button(sel_btns, text="All", width=6, bootstyle="primary-outline", command=lambda: [v.set(True) for _, v in self.chk_vars]).pack(side="left", padx=(0,6))
        tb.Button(sel_btns, text="None", width=6, bootstyle="secondary-outline", command=lambda: [v.set(False) for _, v in self.chk_vars]).pack(side="left")

        # Week panel (center)
        week_panel = tb.Frame(top_container, bootstyle="light", padding=8)
        week_panel.grid(row=1, column=1, sticky="nsew", padx=(8, 8))
        top_container.grid_columnconfigure(1, weight=1)

        # Month and Day side-by-side, editable
        mm_row = tk.Frame(week_panel, bg=PANEL_BG)
        mm_row.pack(anchor="w", pady=(2,6))
        tk.Label(mm_row, text="Month (MM):", bg=PANEL_BG, fg=TEXT_FG).pack(side="left")
        self.month_var = tk.StringVar(value=datetime.now().strftime("%m"))
        self.month_entry = ttk.Entry(mm_row, textvariable=self.month_var, width=6)
        self.month_entry.pack(side="left", padx=(6, 12))

        dd_row = tk.Frame(week_panel, bg=PANEL_BG)
        dd_row.pack(anchor="w", pady=(2,6))
        tk.Label(dd_row, text="Day (DD):", bg=PANEL_BG, fg=TEXT_FG).pack(side="left")
        self.day_var = tk.StringVar(value=datetime.now().strftime("%d"))
        self.day_entry = ttk.Entry(dd_row, textvariable=self.day_var, width=6)
        self.day_entry.pack(side="left", padx=(6, 12))

        tk.Label(week_panel, text="(will search for folders named 'Week MM-DD')", bg=PANEL_BG, fg=TEXT_FG).pack(anchor="w", pady=(6,0))

        # Actions panel (right)
        actions_panel = tb.Frame(top_container, bootstyle="light", padding=8)
        actions_panel.grid(row=1, column=2, sticky="nsew", padx=(8, 4))
        top_container.grid_columnconfigure(2, weight=0)

        self.start_btn = tb.Button(actions_panel, text="Start Merging", width=18, bootstyle="success", command=self.on_start)
        self.start_btn.pack(pady=(2,8))
        tb.Button(actions_panel, text="Exit", width=18, bootstyle="danger", command=self.destroy).pack()

        # Middle large area: status/dashboard
        middle = tk.Frame(root, bg=PANEL_BG)
        middle.pack(fill="both", expand=True, pady=(6, 8))

        # Left: Logo + stats
        left_mid = tk.Frame(middle, bg=PANEL_BG)
        left_mid.pack(side="left", fill="both", expand=True, padx=(0, 8))

        # logo placeholder (you can replace with image)
        logo_frame = tk.Frame(left_mid, width=140, height=120, bg="#e9eef6")
        logo_frame.pack(anchor="nw", padx=8, pady=8)
        logo_frame.pack_propagate(False)
        logo_label = tk.Label(logo_frame, text="LOGO", bg="#e9eef6", fg=TEXT_FG, font=("Segoe UI", 14, "bold"))
        logo_label.pack(expand=True)

        # stats box
        stats_frame = tk.Frame(left_mid, bg=PANEL_BG)
        stats_frame.pack(fill="both", expand=True, padx=8, pady=(6,8))

        tk.Label(stats_frame, text="Status", bg=PANEL_BG, fg=TEXT_FG, font=("Segoe UI", 11, "bold")).pack(anchor="w")
        self.status_vars = {
            "processed": tk.StringVar(value="0"),
            "merged": tk.StringVar(value="0"),
            "warnings": tk.StringVar(value="0"),
            "errors": tk.StringVar(value="0"),
        }

        stats_grid = tk.Frame(stats_frame, bg=PANEL_BG)
        stats_grid.pack(anchor="w", pady=(6,0))
        tk.Label(stats_grid, text="Processed:", bg=PANEL_BG, fg=TEXT_FG).grid(row=0, column=0, sticky="w", padx=4, pady=2)
        tk.Label(stats_grid, textvariable=self.status_vars["processed"], bg=PANEL_BG, fg=TEXT_FG).grid(row=0, column=1, sticky="w", padx=8, pady=2)
        tk.Label(stats_grid, text="Merged:", bg=PANEL_BG, fg=TEXT_FG).grid(row=1, column=0, sticky="w", padx=4, pady=2)
        tk.Label(stats_grid, textvariable=self.status_vars["merged"], bg=PANEL_BG, fg=TEXT_FG).grid(row=1, column=1, sticky="w", padx=8, pady=2)
        tk.Label(stats_grid, text="Warnings:", bg=PANEL_BG, fg=TEXT_FG).grid(row=2, column=0, sticky="w", padx=4, pady=2)
        tk.Label(stats_grid, textvariable=self.status_vars["warnings"], bg=PANEL_BG, fg=TEXT_FG).grid(row=2, column=1, sticky="w", padx=8, pady=2)
        tk.Label(stats_grid, text="Errors:", bg=PANEL_BG, fg=TEXT_FG).grid(row=3, column=0, sticky="w", padx=4, pady=2)
        tk.Label(stats_grid, textvariable=self.status_vars["errors"], bg=PANEL_BG, fg=TEXT_FG).grid(row=3, column=1, sticky="w", padx=8, pady=2)

        # Right: recent activity feed
        right_mid = tk.Frame(middle, bg=PANEL_BG)
        right_mid.pack(side="right", fill="both", expand=True, padx=(8,0))

        tk.Label(right_mid, text="Recent Activity", bg=PANEL_BG, fg=TEXT_FG, font=("Segoe UI", 11, "bold")).pack(anchor="w", padx=4)
        self.activity_text = tk.Text(right_mid, height=10, wrap="word", bg="#f7f9fc", fg=TEXT_FG, relief="flat")
        self.activity_text.pack(fill="both", expand=True, padx=4, pady=(6,8))
        self.activity_text.configure(state="disabled")

        # Progress area (below middle)
        prog_frame = tb.Frame(root, bootstyle="light", padding=8)
        prog_frame.pack(fill="x", pady=(2, 6))

        tk.Label(prog_frame, text="Progress", bg=prog_frame.cget("bg"), fg=TEXT_FG, font=("Segoe UI", 10, "bold")).pack(anchor="w")
        self.progress = tb.Progressbar(prog_frame, mode="determinate")
        self.progress.pack(fill="x", pady=(6, 6))
        self.eta_label = tk.Label(prog_frame, text="0% • ETA: --", bg=prog_frame.cget("bg"), fg=TEXT_FG)
        self.eta_label.pack(anchor="w")

        # Log panel (bottom)
        log_frame = tb.Frame(root, bootstyle="light", padding=8)
        log_frame.pack(fill="both", expand=True, pady=(6, 0))
        tk.Label(log_frame, text="Log", bg=log_frame.cget("bg"), fg=TEXT_FG, font=("Segoe UI", 10, "bold")).pack(anchor="w")
        self.log_text = tk.Text(log_frame, height=8, wrap="word", bg="#ffffff", fg=TEXT_FG, relief="flat")
        self.log_text.pack(fill="both", expand=True, pady=(6,0))
        self.log_text.tag_configure("ok", foreground="#1b8a5a")
        self.log_text.tag_configure("error", foreground="#c62828")
        self.log_text.tag_configure("warn", foreground="#b7791f")
        self.log_text.configure(state="disabled")

        # internal state
        self.pre_scan_info = {}
        self.total_tasks = 0
        self.tasks_done = 0
        self._start_time = None
        self._lock = threading.Lock()

        # initial (fast) scan on background to estimate tasks
        self.after(200, self.background_quick_scan)

    # ---------- quick background scan used to estimate tasks and show initial statuses ----------
    def background_quick_scan(self):
        t = threading.Thread(target=self._quick_scan_thread, daemon=True)
        t.start()

    def _quick_scan_thread(self):
        selected = [name for name, v in self.chk_vars if v.get()]
        week_str = f"{self.month_var.get().zfill(2)}-{self.day_var.get().zfill(2)}"
        pre_scan = {}
        total_tasks = 0
        total_files = 0
        for c in selected:
            client_root = os.path.join(main_folder, c)
            week_path = find_week_folder(client_root, week_str)
            if not week_path:
                pre_scan[c] = {"week": None, "files": [], "tasks": 1}
                total_tasks += 1
            else:
                files = list_raw_files(week_path)
                tasks = 1 + len(files) + 1 + 1  # check + each file + merge + excel
                pre_scan[c] = {"week": week_path, "files": files, "tasks": tasks}
                total_files += len(files)
                total_tasks += tasks

        def ui_update():
            self.pre_scan_info = pre_scan
            self.total_tasks = total_tasks
            # update small status values
            self.status_vars["processed"].set("0")
            self.status_vars["merged"].set("0")
            self.status_vars["warnings"].set("0")
            self.status_vars["errors"].set("0")
            # update recent activity stub
            self._add_activity_line(f"Quick scan: {len(selected)} clients, {total_files} files, est. {total_tasks} tasks")
            self.progress.configure(value=0)
            self.eta_label.configure(text="0% • ETA: --")
        self.after(0, ui_update)

    # ---------- safe UI helpers ----------
    def _add_activity_line(self, text):
        def job():
            try:
                self.activity_text.configure(state="normal")
                ts = datetime.now().strftime("%H:%M:%S")
                self.activity_text.insert("end", f"[{ts}] {text}\n")
                self.activity_text.see("end")
                self.activity_text.configure(state="disabled")
            except Exception:
                pass
        self.activity_text.after(0, job)

    def _log(self, msg, tag="info"):
        # wrapper around logger-like UI write (for use before logger created)
        def job():
            try:
                self.log_text.configure(state="normal")
                if tag in ("error", "ok", "warn"):
                    self.log_text.insert("end", f"{msg}\n", (tag,))
                else:
                    self.log_text.insert("end", f"{msg}\n")
                self.log_text.see("end")
                self.log_text.configure(state="disabled")
            except Exception:
                pass
        self.log_text.after(0, job)

    # ---------- start merging ----------
    def on_start(self):
        selected_clients = [name for name, v in self.chk_vars if v.get()]
        if not selected_clients:
            messagebox.showwarning("Select clients", "Please select at least one client.")
            return
        mm = self.month_var.get().strip()
        dd = self.day_var.get().strip()
        if not (mm.isdigit() and dd.isdigit() and 1 <= int(mm) <= 12 and 1 <= int(dd) <= 31):
            messagebox.showwarning("Invalid date", "Please enter valid numeric Month (MM) and Day (DD).")
            return
        week_str = f"{mm.zfill(2)}-{dd.zfill(2)}"
        self.logger = StepLogger(self.log_text, week_str)

        # ensure we have latest pre_scan info; run quick scan if empty
        if not self.pre_scan_info:
            self.background_quick_scan()
            time.sleep(0.2)

        # compute total tasks from pre_scan
        self.total_tasks = sum(info.get("tasks", 0) for info in self.pre_scan_info.values())
        if self.total_tasks <= 0:
            self.logger.log("Nothing to do for selected clients/date.", "warn")
            return

        # reset counters & UI
        self.tasks_done = 0
        self._start_time = time.time()
        self.status_vars["processed"].set("0")
        self.status_vars["merged"].set("0")
        self.status_vars["warnings"].set("0")
        self.status_vars["errors"].set("0")
        self.progress.configure(value=0)
        self.eta_label.configure(text="0% • ETA: --")
        self._add_activity_line(f"Starting merge for week {week_str}...")
        self.start_btn.configure(state="disabled")

        # start merge in background
        t = threading.Thread(target=self._merge_thread, args=(list(self.pre_scan_info.items()), week_str), daemon=True)
        t.start()

    # ---------- merge worker ----------
    def _merge_thread(self, pre_scan_items, week_str):
        # open excel workbook once
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            self.logger.log(f"Opened Excel: {excel_file}", "ok")
        except Exception as e:
            self.logger.log(f"Cannot open Excel: {e}", "error")
            self._finish(False)
            return

        processed_clients = 0
        merged_count = 0
        warnings = 0
        errors = 0

        for client, info in pre_scan_items:
            # Task: check week folder presence
            try:
                self.logger.log(f"--- Processing client: {client} ---")
                client_root = os.path.join(main_folder, client)
                if not os.path.isdir(client_root):
                    self.logger.log(f"Client folder not found: {client_root}", "error")
                    errors += 1
                    self._increment_task_and_update()
                    continue

                week_path = info.get("week")
                if not week_path:
                    # try resolving dynamically
                    week_path = find_week_folder(client_root, week_str)
                if not week_path:
                    self.logger.log(f"Week folder not found for {client} (week {week_str}).", "warn")
                    warnings += 1
                    self._increment_task_and_update()
                    continue
                self.logger.log(f"Found week folder: {week_path}", "ok")
                self._increment_task_and_update()

                # Task(s): prepare files (convert/rotate). This function logs per-file.
                prepared_list, invoice_final = prepare_files_for_merge(week_path, logger=self.logger)

                # Count prepared raw files => increment per file
                files_raw = info.get("files", [])
                # if quick scan had no file list, update now from folder
                if not files_raw:
                    files_raw = list_raw_files(week_path)
                for _ in files_raw:
                    self._increment_task_and_update()

                # Task: merge prepared files
                if prepared_list:
                    try:
                        merger = PdfMerger()
                        for p in prepared_list:
                            merger.append(p)
                        if invoice_final:
                            out_name = f"{os.path.splitext(os.path.basename(invoice_final))[0]}_.pdf"
                        else:
                            out_name = f"{client}_Week_{week_str}.pdf"
                        out_path = os.path.join(week_path, out_name)
                        merger.write(out_path)
                        merger.close()
                        self.logger.log(f"Merged PDF created: {out_path}", "ok")
                        merged_count += 1
                        self._add_activity_line(f"Merged for {client}: {os.path.basename(out_path)}")
                    except Exception as e:
                        self.logger.log(f"Merge failed for {client}: {e}", "error")
                        errors += 1
                    finally:
                        self._increment_task_and_update()
                else:
                    self.logger.log(f"No files to merge for {client}.", "warn")
                    warnings += 1
                    self._increment_task_and_update()

                # Task: update excel (Col G)
                try:
                    updated = False
                    for r in range(4, ws.max_row + 1):
                        cell_val = ws.cell(row=r, column=2).value
                        if cell_val and str(cell_val).strip() == client:
                            ws.cell(row=r, column=7).value = out_path if prepared_list else ""
                            updated = True
                            break
                    if updated:
                        self.logger.log(f"Excel updated for {client}.", "ok")
                    else:
                        self.logger.log(f"Client '{client}' not found in Excel Col B (rows 4..{ws.max_row}).", "warn")
                        warnings += 1
                except Exception as e:
                    self.logger.log(f"Excel update failed for {client}: {e}", "error")
                    errors += 1
                finally:
                    self._increment_task_and_update()

                processed_clients += 1
                # update status vars on UI thread
                self._safe_set_status(processed=processed_clients, merged=merged_count, warnings=warnings, errors=errors)

            except Exception as e:
                self.logger.log(f"Unexpected failure for {client}: {e}", "error")
                errors += 1
                # attempt to keep progress moving
                self._increment_task_and_update()
                continue

        # Save Excel once at the end
        try:
            wb.save(excel_file)
            self.logger.log("Excel saved.", "ok")
        except Exception as e:
            self.logger.log(f"Excel save error: {e}", "error")

        # finish
        self._finish(True)

    # ---------- helpers for thread-safe UI updates ----------
    def _increment_task_and_update(self):
        with self._lock:
            self.tasks_done += 1
            pct = int(self.tasks_done / max(1, self.total_tasks) * 100)
            elapsed = time.time() - (self._start_time or time.time())
            avg = elapsed / max(1, self.tasks_done)
            remaining_seconds = int(avg * max(0, (self.total_tasks - self.tasks_done)))
            # schedule UI updates
            try:
                self.progress.after(0, lambda v=pct: self.progress.configure(value=v))
                self.eta_label.after(0, lambda t=f"{pct}% • ETA: {remaining_seconds}s": self.eta_label.configure(text=t))
            except Exception:
                pass

    def _safe_set_status(self, processed=None, merged=None, warnings=None, errors=None):
        def job():
            try:
                if processed is not None:
                    self.status_vars["processed"].set(str(processed))
                if merged is not None:
                    self.status_vars["merged"].set(str(merged))
                if warnings is not None:
                    self.status_vars["warnings"].set(str(warnings))
                if errors is not None:
                    self.status_vars["errors"].set(str(errors))
            except Exception:
                pass
        self.after(0, job)

    def _finish(self, success: bool):
        # short standard system beep (not deep)
        try:
            winsound.MessageBeep(winsound.MB_OK)
        except Exception:
            try:
                winsound.Beep(800, 200)
            except Exception:
                pass
        # final status update
        self._safe_set_status()  # no-op if not provided
        # ensure progress UI reports done
        try:
            self.progress.after(0, lambda: self.progress.configure(value=100))
            self.eta_label.after(0, lambda: self.eta_label.configure(text="Done ✅"))
            self.start_btn.after(0, lambda: self.start_btn.configure(state="normal"))
        except Exception:
            pass
        self._add_activity_line("Merging process finished.")

# ============ Run ============
if __name__ == "__main__":
    app = App()
    app.title(APP_TITLE)
    # keep a subtle outer bg so the window feels branded
    app.configure(bg=THEME_BG)
    app.mainloop()
