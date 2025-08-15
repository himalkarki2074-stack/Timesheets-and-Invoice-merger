# invoice_timesheets_compiler_final.py
import os
import time
import threading
import winsound
from datetime import datetime

# UI
import tkinter as tk
from tkinter import messagebox, ttk
import ttkbootstrap as tb
from ttkbootstrap.constants import *

# File handling / conversions
from PyPDF2 import PdfMerger, PdfReader, PdfWriter
from PIL import Image, ImageOps
from docx2pdf import convert
import openpyxl

# ===================== CONFIG =====================
main_folder = r"O:\ApTask\TDrive\FinTech LLC\Invoices\2025\Monthly"
excel_file  = r"C:\Users\HimalK\OneDrive - APTASK\Desktop\Aptask\Payroll\Automated email sheet\Emailexcel.xlsx"
log_folder  = r"C:\Users\HimalK\OneDrive - APTASK\Desktop\Aptask\Payroll\Timesheet_Invoice Merger\Logfile"
os.makedirs(log_folder, exist_ok=True)

clients_list = [
    "Aquila Energy", "BDR", "B Squared", "CFAIS", "Data Specialist",
    "HTS Workforce", "Schultz Controls", "Security 101", "VFS Fire", "Western Audio"
]

APP_TITLE = "Invoice and Timesheets Compiler"
THEME_BG  = "#2b6cb0"   # medium blue background
PANEL_BG  = "#ffffff"   # white panels for contrast
TEXT_FG   = "#000000"   # black text as requested
ACCENT    = "#1e90ff"
# ===================================================


# ---------- Logger ----------
class StepLogger:
    def __init__(self, text_widget, week_str):
        self.text_widget = text_widget
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_week = week_str.replace("/", "-")
        self.path = os.path.join(log_folder, f"Log_Week_{safe_week}_{ts}.txt")

    def _write_file(self, line: str):
        try:
            with open(self.path, "a", encoding="utf-8") as f:
                f.write(line + "\n")
        except Exception:
            pass

    def log(self, msg, tag="info"):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        line = f"[{timestamp}] {msg}"
        self.text_widget.configure(state="normal")
        if tag == "error":
            self.text_widget.insert("end", line + "\n", ("error",))
        elif tag == "ok":
            self.text_widget.insert("end", line + "\n", ("ok",))
        elif tag == "warn":
            self.text_widget.insert("end", line + "\n", ("warn",))
        else:
            self.text_widget.insert("end", line + "\n")
        self.text_widget.see("end")
        self.text_widget.configure(state="disabled")
        self._write_file(line)


# ---------- Helpers ----------
def rotate_pdf_if_needed(src_pdf, dst_pdf):
    reader = PdfReader(src_pdf)
    writer = PdfWriter()
    for page in reader.pages:
        try:
            w = float(page.mediabox.width)
            h = float(page.mediabox.height)
            if w > h:
                page.rotate(90)
        except Exception:
            pass
        writer.add_page(page)
    with open(dst_pdf, "wb") as f:
        writer.write(f)


def image_to_pdf(image_path, out_pdf):
    image = Image.open(image_path)
    image = ImageOps.exif_transpose(image)
    image.convert("RGB").save(out_pdf)


def word_to_rotated_pdf(doc_path, out_pdf):
    temp_pdf = out_pdf.replace(".pdf", "_raw.pdf")
    convert(doc_path, temp_pdf)  # requires Word installed on Windows
    rotate_pdf_if_needed(temp_pdf, out_pdf)
    try:
        os.remove(temp_pdf)
    except Exception:
        pass


def find_week_folder(client_root, week_str):
    target = f"Week {week_str}"
    if not os.path.isdir(client_root):
        return None
    for month_name in sorted(os.listdir(client_root)):
        month_path = os.path.join(client_root, month_name)
        if not os.path.isdir(month_path):
            continue
        week_path = os.path.join(month_path, target)
        if os.path.isdir(week_path):
            return week_path
    return None


def list_raw_files(folder):
    files = []
    if not os.path.isdir(folder):
        return files
    for name in sorted(os.listdir(folder)):
        path = os.path.join(folder, name)
        if not os.path.isfile(path):
            continue
        ext = os.path.splitext(name)[1].lower()
        if ext in (".pdf", ".jpg", ".jpeg", ".png", ".docx", ".doc"):
            files.append(path)
    return files


# ---------- App ----------
class App(tb.Window):
    def __init__(self):
        super().__init__(themename="flatly")
        self.title(APP_TITLE)
        self.geometry("1100x920")
        self.resizable(True, True)
        self.configure(bg=THEME_BG)

        # Outer
        outer = tb.Frame(self, bootstyle="secondary", padding=(10, 10))
        outer.pack(fill="both", expand=True, padx=10, pady=10)

        # Header (white title stripe so black text visible)
        header_frame = tk.Frame(outer, bg=PANEL_BG)
        header_frame.pack(fill="x", pady=(0, 8))
        header_label = tk.Label(header_frame, text=APP_TITLE, font=("Segoe UI", 20, "bold"),
                                bg=PANEL_BG, fg=TEXT_FG)
        header_label.pack(padx=8, pady=8, anchor="w")

        # Top row
        top_row = tb.Frame(outer)
        top_row.pack(fill="x", pady=(0, 10))

        # Clients label (explicit black) + panel
        lbl_clients = tk.Label(top_row, text="Clients", bg=THEME_BG, fg=TEXT_FG, font=("Segoe UI", 10, "bold"))
        lbl_clients.pack(side="left", anchor="n", padx=(6, 0))
        clients_panel = tb.Frame(top_row, bootstyle="light", padding=8)
        clients_panel.pack(side="left", fill="y", padx=(6, 12))

        # use ttkbootstrap checkbuttons for the nicer checkmark look, ensure text black
        self.chk_vars = []
        for c in clients_list:
            var = tk.BooleanVar(value=True)
            # style tweak: use bootstyle but create with a small label font and black text
            cb = tb.Checkbutton(clients_panel, text=c, variable=var, bootstyle="round-toggle")
            # try forcing label color to black (ttk sometimes ignores fg param):
            try:
                cb.configure(foreground=TEXT_FG)
            except Exception:
                pass
            cb.pack(anchor="w", pady=2, padx=2)
            self.chk_vars.append((c, var))

        # small select/none
        btns = tb.Frame(clients_panel)
        btns.pack(anchor="w", pady=(6, 0))
        tb.Button(btns, text="All", width=6, bootstyle="primary-outline",
                  command=lambda: [v.set(True) for _, v in self.chk_vars]).pack(side="left", padx=(0, 6))
        tb.Button(btns, text="None", width=6, bootstyle="secondary-outline",
                  command=lambda: [v.set(False) for _, v in self.chk_vars]).pack(side="left")

        # Date label + panel
        lbl_date = tk.Label(top_row, text="Week Ending (MM-DD)", bg=THEME_BG, fg=TEXT_FG, font=("Segoe UI", 10, "bold"))
        lbl_date.pack(side="left", anchor="n", padx=(6, 0))
        date_panel = tb.Frame(top_row, bootstyle="light", padding=8)
        date_panel.pack(side="left", fill="both", expand=True, padx=(6, 12))

        # Month & day comboboxes
        mrow = tb.Frame(date_panel)
        mrow.pack(anchor="w", pady=2, fill="x")
        tk.Label(mrow, text="Month (MM):", bg=PANEL_BG, fg=TEXT_FG).pack(side="left")
        self.month_var = tk.StringVar(value=datetime.now().strftime("%m"))
        self.month_combo = ttk.Combobox(mrow, textvariable=self.month_var, width=6, state="readonly")
        self.month_combo['values'] = [f"{i:02d}" for i in range(1, 13)]
        self.month_combo.pack(side="left", padx=(6, 0))

        drow = tb.Frame(date_panel)
        drow.pack(anchor="w", pady=6, fill="x")
        tk.Label(drow, text="Day (DD):", bg=PANEL_BG, fg=TEXT_FG).pack(side="left")
        self.day_var = tk.StringVar(value=datetime.now().strftime("%d"))
        self.day_combo = ttk.Combobox(drow, textvariable=self.day_var, width=6, state="readonly")
        self.day_combo['values'] = [f"{i:02d}" for i in range(1, 32)]
        self.day_combo.pack(side="left", padx=(6, 0))

        tk.Label(date_panel, text="(scans for folders named 'Week MM-DD')", bg=PANEL_BG, fg=TEXT_FG).pack(anchor="w", pady=(6, 0))

        # Actions label + panel
        lbl_actions = tk.Label(top_row, text="Actions", bg=THEME_BG, fg=TEXT_FG, font=("Segoe UI", 10, "bold"))
        lbl_actions.pack(side="right", anchor="n", padx=(0, 8))
        actions = tb.Frame(top_row, bootstyle="light", padding=8)
        actions.pack(side="right", fill="y")

        self.start_btn = tb.Button(actions, text="Start Merging", width=16, bootstyle="success", command=self.on_start)
        self.start_btn.pack(pady=(2, 6))
        tb.Button(actions, text="Exit", width=16, bootstyle="danger", command=self.destroy).pack()

        # Middle area: Progress & Log & Stats
        middle = tb.Frame(outer)
        middle.pack(fill="both", expand=True)

        # Left column: Progress & Log
        left_col = tb.Frame(middle)
        left_col.pack(side="left", fill="both", expand=True, padx=(0, 8))

        prog_panel = tb.Labelframe(left_col, text="Progress", bootstyle="light", padding=10)
        prog_panel.pack(fill="x", pady=(0, 8))
        self.progress = tb.Progressbar(prog_panel, mode="determinate", length=700)
        self.progress.pack(fill="x", pady=(0, 6))
        try:
            self.progress.configure(maximum=100)
        except Exception:
            pass
        self.eta_label = tk.Label(prog_panel, text="0% • ETA: --", bg=PANEL_BG, fg=TEXT_FG)
        self.eta_label.pack(anchor="w")

        log_panel = tb.Labelframe(left_col, text="Log", bootstyle="light", padding=10)
        log_panel.pack(fill="both", expand=True)
        self.log_text = tk.Text(log_panel, height=18, wrap="word", bg=PANEL_BG, fg=TEXT_FG, relief="flat", font=("Segoe UI", 10))
        self.log_text.pack(fill="both", expand=True)
        self.log_text.tag_configure("ok", foreground="#1b8a5a")
        self.log_text.tag_configure("error", foreground="#c62828")
        self.log_text.tag_configure("warn", foreground="#b7791f")
        self.log_text.configure(state="disabled")

        # Right column: Stats & Queue (fill middle space)
        right_col = tb.Labelframe(middle, text="Stats & Queue", bootstyle="light", padding=10)
        right_col.pack(side="right", fill="y", ipadx=8, ipady=8)

        tk.Label(right_col, text="Selected clients:", bg=PANEL_BG, fg=TEXT_FG).pack(anchor="w")
        self.lbl_sel_clients = tk.Label(right_col, text="0", bg=PANEL_BG, fg=TEXT_FG)
        self.lbl_sel_clients.pack(anchor="w", pady=(0, 8))

        tk.Label(right_col, text="Total files queued:", bg=PANEL_BG, fg=TEXT_FG).pack(anchor="w")
        self.lbl_total_files = tk.Label(right_col, text="0", bg=PANEL_BG, fg=TEXT_FG)
        self.lbl_total_files.pack(anchor="w", pady=(0, 8))

        tk.Label(right_col, text="Estimated tasks:", bg=PANEL_BG, fg=TEXT_FG).pack(anchor="w")
        self.lbl_total_tasks = tk.Label(right_col, text="0", bg=PANEL_BG, fg=TEXT_FG)
        self.lbl_total_tasks.pack(anchor="w", pady=(0, 8))

        tb.Separator(right_col).pack(fill="x", pady=6)

        tk.Label(right_col, text="Client queue preview:", bg=PANEL_BG, fg=TEXT_FG).pack(anchor="w")
        self.client_select = ttk.Combobox(right_col, state="readonly")
        self.client_select.pack(fill="x", pady=(4, 6))
        self.client_select.bind("<<ComboboxSelected>>", self.on_client_selected)

        # queue list with scrollbar (ensures not hidden)
        self.queue_frame = tk.Frame(right_col)
        self.queue_frame.pack(fill="both", expand=True)
        self.queue_scroll = tk.Scrollbar(self.queue_frame, orient="vertical")
        self.queue_list = tk.Listbox(self.queue_frame, width=40, height=12, bg="white", fg=TEXT_FG, yscrollcommand=self.queue_scroll.set)
        self.queue_scroll.config(command=self.queue_list.yview)
        self.queue_list.pack(side="left", fill="both", expand=True)
        self.queue_scroll.pack(side="right", fill="y")

        tb.Button(right_col, text="Refresh Scan", bootstyle="outline-primary", command=self.start_background_scan).pack(pady=(8, 0))

        # internal
        self.pre_scan_info = {}
        self.total_tasks = 0
        self.tasks_done = 0
        self.scan_thread = None

        # initial scan in background for responsive startup
        self.start_background_scan()

    # ---------- Scanning runs in background ----------
    def start_background_scan(self):
        # don't start another if already running
        if self.scan_thread and self.scan_thread.is_alive():
            return
        self.log_text.configure(state="normal")
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.log_text.insert("end", f"[{ts}] Starting background scan...\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")
        self.scan_thread = threading.Thread(target=self.refresh_scan, daemon=True)
        self.scan_thread.start()

    def refresh_scan(self):
        sel_clients = [name for name, v in self.chk_vars if v.get()]
        pre_scan = {}
        total_files = 0
        total_tasks = 0
        week_str = f"{self.month_combo.get()}-{self.day_combo.get()}"
        # iterate quickly (no conversions)
        for c in sel_clients:
            client_root = os.path.join(main_folder, c)
            week_path = find_week_folder(client_root, week_str)
            if not week_path:
                pre_scan[c] = {"week": None, "files": [], "tasks": 1}
                total_tasks += 1
            else:
                raw_files = list_raw_files(week_path)
                tasks = 1 + len(raw_files) + 1 + 1
                pre_scan[c] = {"week": week_path, "files": raw_files, "tasks": tasks}
                total_files += len(raw_files)
                total_tasks += tasks

        # update UI on main thread
        def update_ui():
            self.pre_scan_info = pre_scan
            self.lbl_sel_clients.configure(text=str(len(sel_clients)))
            self.lbl_total_files.configure(text=str(total_files))
            self.lbl_total_tasks.configure(text=str(total_tasks))
            self.total_tasks = total_tasks
            all_clients = list(self.pre_scan_info.keys())
            if not all_clients:
                all_clients = sel_clients
            self.client_select['values'] = all_clients
            if all_clients:
                try:
                    self.client_select.set(all_clients[0])
                except Exception:
                    pass
                self.update_queue_list(all_clients[0])
            else:
                self.client_select.set("")
                self.queue_list.delete(0, "end")
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.log_text.configure(state="normal")
            self.log_text.insert("end", f"[{ts}] Scan complete: {len(sel_clients)} clients, {total_files} files, {total_tasks} estimated tasks\n")
            self.log_text.see("end")
            self.log_text.configure(state="disabled")

        self.after(0, update_ui)

    def update_queue_list(self, client):
        self.queue_list.delete(0, "end")
        info = self.pre_scan_info.get(client, {})
        files = info.get("files", [])
        if not files:
            self.queue_list.insert("end", "(no files found)")
            return
        for f in files:
            self.queue_list.insert("end", os.path.basename(f))

    def on_client_selected(self, event):
        c = self.client_select.get()
        if c:
            self.update_queue_list(c)

    # ---------- Start ----------
    def on_start(self):
        sel_clients = [name for name, v in self.chk_vars if v.get()]
        if not sel_clients:
            messagebox.showwarning("Select clients", "Please select at least one client.")
            return
        mm = self.month_combo.get()
        dd = self.day_combo.get()
        if not (mm and dd):
            messagebox.showwarning("Select date", "Please choose month and day (MM-DD).")
            return

        week_str = f"{mm}-{dd}"
        self.logger = StepLogger(self.log_text, week_str)
        # ensure latest scan results (run synchronously if background scan still running)
        if self.scan_thread and self.scan_thread.is_alive():
            # wait a short time for it to finish
            self.scan_thread.join(timeout=1)

        # compute tasks
        self.total_tasks = sum(info.get("tasks", 0) for info in self.pre_scan_info.values())
        if self.total_tasks <= 0:
            self.logger.log("Nothing to do for selected clients and date.", "warn")
            return
        self.tasks_done = 0
        try:
            self.progress.configure(value=0)
        except Exception:
            pass
        self.eta_label.configure(text="0% • ETA: --")
        self.start_btn.configure(state="disabled")
        self._start_time = time.time()
        t = threading.Thread(target=self._run_merge, args=(list(self.pre_scan_info.items()), week_str), daemon=True)
        t.start()

    # ---------- run merging thread (granular updates) ----------
    def _run_merge(self, pre_scan_items, week_str):
        # open excel
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            self.logger.log(f"Opened Excel: {excel_file}", "ok")
        except Exception as e:
            self.logger.log(f"Cannot open Excel: {e}", "error")
            self._finish(False)
            return

        for client, info in pre_scan_items:
            try:
                self.logger.log(f"--- Processing client: {client} ---")
                client_root = os.path.join(main_folder, client)
                week_path = info.get("week")
                # Task: check week folder
                if not week_path:
                    self.logger.log(f"Week folder not found for {client} -> skipping", "warn")
                    self._increment_task_and_update()
                    continue
                else:
                    self.logger.log(f"Found week folder: {week_path}", "ok")
                    self._increment_task_and_update()

                # process files
                raw_files = info.get("files", [])
                processed_pdfs = []
                for fpath in raw_files:
                    try:
                        root, ext = os.path.splitext(os.path.basename(fpath))
                        ext = ext.lower()
                        if ext == ".pdf":
                            out_pdf = os.path.join(week_path, f"{root}_rotated.pdf")
                            rotate_pdf_if_needed(fpath, out_pdf)
                            processed_pdfs.append(out_pdf)
                            self.logger.log(f"Processed PDF: {os.path.basename(fpath)}", "ok")
                        elif ext in (".jpg", ".jpeg", ".png"):
                            out_pdf = os.path.join(week_path, f"{root}.pdf")
                            image_to_pdf(fpath, out_pdf)
                            processed_pdfs.append(out_pdf)
                            self.logger.log(f"Processed Image: {os.path.basename(fpath)}", "ok")
                        elif ext in (".docx", ".doc"):
                            out_pdf = os.path.join(week_path, f"{root}.pdf")
                            word_to_rotated_pdf(fpath, out_pdf)
                            processed_pdfs.append(out_pdf)
                            self.logger.log(f"Processed Word: {os.path.basename(fpath)}", "ok")
                        else:
                            self.logger.log(f"Skipped unsupported file: {os.path.basename(fpath)}", "warn")
                    except Exception as e:
                        self.logger.log(f"Error processing {os.path.basename(fpath)}: {e}", "error")
                    finally:
                        self._increment_task_and_update()

                # find invoice
                invoice_pdf = None
                for name in sorted(os.listdir(week_path)):
                    if name.lower().endswith(".pdf") and "invoice" in name.lower():
                        invoice_pdf = os.path.join(week_path, name)
                        break

                # merge
                try:
                    merge_list = []
                    if invoice_pdf:
                        merge_list.append(invoice_pdf)
                    merge_list.extend(processed_pdfs)
                    if merge_list:
                        merger = PdfMerger()
                        for p in merge_list:
                            merger.append(p)
                        if invoice_pdf:
                            inv_base = os.path.splitext(os.path.basename(invoice_pdf))[0]
                            out_name = f"{inv_base}_.pdf"
                        else:
                            out_name = f"{client}_Week_{week_str}.pdf"
                        out_path = os.path.join(week_path, out_name)
                        merger.write(out_path)
                        merger.close()
                        self.logger.log(f"Merged PDF created: {out_path}", "ok")
                    else:
                        out_path = ""
                        self.logger.log("No files to merge for this client.", "warn")
                except Exception as e:
                    out_path = ""
                    self.logger.log(f"Merge failed for {client}: {e}", "error")
                finally:
                    self._increment_task_and_update()

                # update excel
                try:
                    start_row, last_row = 4, ws.max_row
                    updated = False
                    for r in range(start_row, last_row + 1):
                        name_cell = ws.cell(row=r, column=2).value
                        if str(name_cell).strip() == client:
                            ws.cell(row=r, column=7).value = out_path
                            updated = True
                            break
                    if updated:
                        self.logger.log(f"Excel updated for {client} (Col G).", "ok")
                    else:
                        self.logger.log(f"Client '{client}' not found in Excel Col B (rows 4..{last_row}).", "warn")
                except Exception as e:
                    self.logger.log(f"Excel update failed for {client}: {e}", "error")
                finally:
                    self._increment_task_and_update()

            except Exception as e:
                self.logger.log(f"Client '{client}' failed: {e}", "error")
                continue

        # save excel
        try:
            wb.save(excel_file)
            self.logger.log("Excel saved.", "ok")
        except Exception as e:
            self.logger.log(f"Excel save error: {e}", "error")

        self._finish(True)

    def _increment_task_and_update(self):
        self.tasks_done += 1
        pct = int(self.tasks_done / max(1, self.total_tasks) * 100)
        elapsed = time.time() - getattr(self, "_start_time", time.time())
        avg = elapsed / max(1, self.tasks_done)
        remaining_seconds = int(avg * max(0, (self.total_tasks - self.tasks_done)))
        # schedule UI updates
        self.progress.after(0, lambda v=pct: self.progress.configure(value=v))
        self.eta_label.after(0, lambda t=f"{pct}% • ETA: {remaining_seconds}s": self.eta_label.configure(text=t))

    def _finish(self, success: bool):
        # short system notification (no deep tone)
        try:
            winsound.MessageBeep(winsound.MB_OK)
        except Exception:
            try:
                winsound.Beep(750, 200)
            except Exception:
                pass
        msg = "Merging Completed ✅" if success else "Completed with Errors ⚠️"
        self.logger.log(msg, "ok" if success else "warn")
        self.start_btn.after(0, lambda: self.start_btn.configure(state="normal"))
        self.eta_label.after(0, lambda: self.eta_label.configure(text=("Done ✅" if success else "Done ⚠️")))


# ---------- Run ----------
if __name__ == "__main__":
    app = App()
    app.title(APP_TITLE)
    app.configure(bg=THEME_BG)
    app.mainloop()
